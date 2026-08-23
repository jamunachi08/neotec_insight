"""Tests for utils/cash_flow_import.py. Uses a small synthetic workbook
built with real openpyxl, matching the shape verified against the actual
customer file this feature was built from — trailing spaces on headers,
header row not at row 1, a mix of matched and unmatched categories — rather
than depending on that uploaded file being present wherever this test runs.
"""

from __future__ import annotations

import io
import sys
import unittest
from pathlib import Path

import openpyxl

APP_ROOT = Path(__file__).resolve().parents[1] / "neotec_insight" / "utils"
sys.path.insert(0, str(APP_ROOT))

import cash_flow_import as imp  # noqa: E402


def _build_workbook(header_row=4, include_trailing_spaces=True, sheet_name="Data") -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = {
        1: "Month", 2: "Month", 3: "Posting Date", 4: "Account",
        5: "Transaction Type " if include_trailing_spaces else "Transaction Type",
        6: "Class", 7: "New Class", 8: "Debit (SAR)", 9: "Credit (SAR)", 10: "Balance (SAR)",
        11: "Voucher Type", 12: "Voucher No", 13: "Against Account",
        14: "Project", 15: "Cost Center", 16: "Remarks",
    }
    for col, label in headers.items():
        ws.cell(row=header_row, column=col, value=label)

    data_rows = [
        # month, month, date, account, tt, class, new_class, debit, credit, balance,
        # vtype, vno, against, project, cc, remarks
        (1, 1, "2026-01-01", "Riyad Bank", "Cash In", "Direct", "Collection", 6900, 0, 6900,
         "Payment Entry", "ACC-PAY-2026-00001", "Tazweed Co", None, "Book Keeping",
         "Received from Tazweed for invoice"),
        (1, 1, "2026-01-05", "Riyad Bank", "Cash Out", "Direct", "GOSI Payment", 0, 11400, -11400,
         "Journal Entry", "ACC-JV-2026-00002", "GOSI", None, "Audit", "GOSI payment for Jan"),
        (1, 1, "2026-01-10", "Riyad Bank", "Cash Out", "Overhead", "A Category Not In The App",
         0, 500, -500, "Payment Entry", "ACC-PAY-2026-00003", "Someone", None, "Admin",
         "Something not yet configured"),
    ]
    for r, row in enumerate(data_rows, start=header_row + 1):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestParseClassifiedHistorySheet(unittest.TestCase):
    def test_finds_header_row_not_at_row_1(self):
        result = imp.parse_classified_history_sheet(_build_workbook(header_row=4))
        self.assertEqual(result["header_row"], 4)

    def test_finds_header_row_at_a_different_position(self):
        """The header row is discovered by content, not assumed at a fixed
        position — a workbook with the header at row 7 must work the same
        as one with it at row 4."""
        result = imp.parse_classified_history_sheet(_build_workbook(header_row=7))
        self.assertEqual(result["header_row"], 7)
        self.assertEqual(len(result["rows"]), 3)

    def test_trailing_spaces_on_headers_do_not_break_matching(self):
        result = imp.parse_classified_history_sheet(_build_workbook(include_trailing_spaces=True))
        self.assertIn("transaction_type", result["columns_found"])

    def test_extracts_all_data_rows(self):
        result = imp.parse_classified_history_sheet(_build_workbook())
        self.assertEqual(len(result["rows"]), 3)

    def test_row_fields_are_correctly_mapped(self):
        result = imp.parse_classified_history_sheet(_build_workbook())
        row = result["rows"][0]
        self.assertEqual(row["new_class"], "Collection")
        self.assertEqual(row["voucher_type"], "Payment Entry")
        self.assertEqual(row["voucher_no"], "ACC-PAY-2026-00001")
        self.assertEqual(row["remarks"], "Received from Tazweed for invoice")
        self.assertEqual(row["debit"], 6900)

    def test_blank_new_class_row_is_skipped_not_erroring(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        for col, label in [(11, "Voucher Type"), (12, "Voucher No"), (7, "New Class"), (16, "Remarks")]:
            ws.cell(row=4, column=col, value=label)
        ws.cell(row=5, column=11, value="Payment Entry")
        ws.cell(row=5, column=12, value="ACC-PAY-1")
        # New Class deliberately left blank on this row
        buf = io.BytesIO()
        wb.save(buf)
        result = imp.parse_classified_history_sheet(buf.getvalue())
        self.assertEqual(result["rows"], [])

    def test_no_header_row_found_returns_empty_with_a_warning(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Not a real header sheet")
        buf = io.BytesIO()
        wb.save(buf)
        result = imp.parse_classified_history_sheet(buf.getvalue())
        self.assertEqual(result["rows"], [])
        self.assertTrue(result["warnings"])

    def test_explicit_sheet_name_is_used_when_present(self):
        result = imp.parse_classified_history_sheet(_build_workbook(sheet_name="Data"), sheet_name="Data")
        self.assertEqual(result["sheet_used"], "Data")

    def test_missing_explicit_sheet_name_falls_back_with_a_warning(self):
        result = imp.parse_classified_history_sheet(_build_workbook(sheet_name="Data"), sheet_name="DoesNotExist")
        self.assertEqual(result["sheet_used"], "Data")
        self.assertTrue(any("DoesNotExist" in w for w in result["warnings"]))


class TestMatchLinesToRows(unittest.TestCase):
    def setUp(self):
        self.rows = imp.parse_classified_history_sheet(_build_workbook())["rows"]
        self.lines = [
            {"name": "line-collection", "label": "Collection"},
            {"name": "line-gosi", "label": "GOSI Payment"},
        ]

    def test_matched_rows_get_the_real_line_name(self):
        result = imp.match_lines_to_rows(self.rows, self.lines)
        self.assertEqual(result["matched_count"], 2)
        matched_lines = {r["new_class"]: r["line"] for r in result["matched"]}
        self.assertEqual(matched_lines["Collection"], "line-collection")
        self.assertEqual(matched_lines["GOSI Payment"], "line-gosi")

    def test_unmatched_category_is_reported_by_name_and_count_not_dropped_silently(self):
        result = imp.match_lines_to_rows(self.rows, self.lines)
        self.assertEqual(result["unmatched_count"], 1)
        self.assertIn("A Category Not In The App", result["unmatched_labels"])
        self.assertEqual(result["unmatched_labels"]["A Category Not In The App"], 1)

    def test_matching_is_case_and_whitespace_insensitive(self):
        rows = [{"new_class": "  gosi payment  ", "voucher_type": "Payment Entry",
                 "voucher_no": "V1", "remarks": "", "debit": 0, "credit": 100,
                 "against_account": "", "cost_center": "", "project": ""}]
        result = imp.match_lines_to_rows(rows, self.lines)
        self.assertEqual(result["matched_count"], 1)
        self.assertEqual(result["matched"][0]["line"], "line-gosi")

    def test_no_lines_available_means_everything_is_unmatched(self):
        result = imp.match_lines_to_rows(self.rows, [])
        self.assertEqual(result["matched_count"], 0)
        self.assertEqual(result["unmatched_count"], 3)


if __name__ == "__main__":
    unittest.main(verbosity=2)
