"""Cash Flow Forecast — historical import engine.

Same isolation boundary as the rest of Cash Flow Forecast — no import from
utils/map_importer.py or the P&L engine, even though the shape of this
feature (upload an .xlsx, parse with openpyxl, apply to app data) matches
that existing feature's own established pattern. Following the pattern is
not the same as sharing the code.

Three pure concerns, all testable without a live site:
  1. parse_classified_history_sheet — read the "Data" sheet shape a real
     customer's manual process produces (Month/Posting Date/Account/
     Transaction Type/Class/New Class/Debit/Credit/Balance/Voucher Type/
     Voucher No/Against Account/Project/Cost Center/Remarks) into plain rows.
  2. match_lines_to_rows — resolve each row's "New Class" label to a real
     Insight Cash Flow Line, or report it as unmatched so the user knows
     exactly which categories still need a Line created before import can
     claim them.
  3. parse_statement_template_sheet — read the customer's OWN month-by-month
     Budget/Actual statement layout (the shape they've been maintaining by
     hand) and turn it into a Line list + Budget grid — the import that
     actually solves "no Lines exist yet, so nothing in the transaction
     import can match": this one creates the Lines themselves, plus every
     Budget figure already on the sheet, in one pass.
"""

from __future__ import annotations

import io

import openpyxl


_MONTH_PREFIXES = {"jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
                   "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12}


def _norm(v) -> str:
    return str(v or "").strip()


def parse_statement_template_sheet(file_bytes: bytes, sheet_name: str | None = None,
                                   fiscal_year: int | None = None) -> dict:
    """Returns {lines, warnings, sheet_used, fiscal_year_guess}.

    `lines`: [{label, direction, section, budgets: {month_number: amount}}].
    direction is "Cash Out" until a section header's own label contains
    "collection" or "cash in" (case-insensitive), matching the two-block
    shape verified against the real customer template this was built from
    (a "Cash out Item" block, then a "Collection/Department" block) — every
    row in a matched-"cash in" section becomes its own Cash In line, one per
    department, rather than assuming a single shared Collection line the
    way Account Binding's cost-centre dimension does; that's a real
    simplification worth knowing about, not a hidden equivalence, since the
    two produce different Line Setup outcomes.

    A header row is recognized by an "SL." cell (case/period-insensitive);
    the label column is the one immediately after it. Real month columns
    are recognized by their header text's first three letters matching a
    month name ("Jan", "March", "April", ...) — quarter/total subtotal
    columns interspersed in the real template ("Q1", "Q2", "Total") do not
    match any three-letter month prefix and are skipped automatically,
    without needing to special-case their exact labels."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    warnings: list[str] = []

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_used = sheet_name
    else:
        ws = wb.worksheets[0]
        sheet_used = ws.title
        if sheet_name:
            warnings.append(f"Sheet '{sheet_name}' not found — used '{sheet_used}' instead.")

    def _find_header(start_row: int) -> tuple[int, int] | None:
        """Returns (header_row, label_col) for the first 'SL.' header found
        at or after start_row, or None if there isn't one."""
        for r in range(start_row, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = _norm(ws.cell(row=r, column=c).value).lower().rstrip(".")
                if cell == "sl":
                    return r, c + 1
        return None

    first = _find_header(1)
    if not first:
        return {"lines": [], "warnings": ["Could not find a header row with an 'SL.' column."],
                "sheet_used": sheet_used, "fiscal_year_guess": fiscal_year}
    header_row, label_col = first
    # The first header row's own label ("Cash out Item" in the real
    # template) is the starting section — captured here explicitly, since
    # the row-scan loop below only updates current_section when it meets a
    # LATER header row; without this, every row before the second section
    # would report a blank section, only ever falling back to the raw
    # direction name.
    initial_section = _norm(ws.cell(row=header_row, column=label_col).value)

    def _month_columns(row: int) -> dict[int, int]:
        """{month_number: budget_column} — actual is assumed to sit in the
        very next column, matching the real template's Budget/Actual pairing
        confirmed at every one of its 12 real month columns."""
        out = {}
        for c in range(label_col + 1, ws.max_column + 1):
            label = _norm(ws.cell(row=row, column=c).value).lower()
            month_no = _MONTH_PREFIXES.get(label[:3])
            if month_no and month_no not in out:
                out[month_no] = c
        return out

    fy_guess = fiscal_year
    if fy_guess is None:
        for r in range(1, header_row):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if hasattr(v, "year"):
                    fy_guess = v.year
                    break
            if fy_guess:
                break

    lines: list[dict] = []
    month_cols = _month_columns(header_row)
    if not month_cols:
        warnings.append(f"No recognizable month columns found in the header row at row {header_row}.")

    current_section = initial_section
    current_direction = "Cash In" if ("collection" in initial_section.lower()
                                      or "cash in" in initial_section.lower()) else "Cash Out"
    r = header_row + 1
    while r <= ws.max_row:
        sl_cell = _norm(ws.cell(row=r, column=label_col - 1).value)
        label = _norm(ws.cell(row=r, column=label_col).value)

        if sl_cell.lower().rstrip(".") == "sl":
            # A new section header — its own label is the section name, and
            # everything below it uses a freshly-scanned set of month
            # columns (defensive: nothing requires column layout to repeat
            # identically between sections, even though it does in the real
            # template this was built from).
            current_section = label
            current_direction = "Cash In" if ("collection" in label.lower()
                                              or "cash in" in label.lower()) else "Cash Out"
            month_cols = _month_columns(r) or month_cols
            r += 1
            continue

        if not label or not sl_cell:
            r += 1
            continue

        budgets = {}
        for month_no, bcol in month_cols.items():
            bval = ws.cell(row=r, column=bcol).value
            if bval not in (None, ""):
                try:
                    budgets[month_no] = float(bval)
                except (TypeError, ValueError):
                    pass
        lines.append({
            "label": label, "direction": current_direction,
            "section": current_section or current_direction,
            "budgets": budgets,
        })
        r += 1

    return {"lines": lines, "warnings": warnings, "sheet_used": sheet_used,
            "fiscal_year_guess": fy_guess}


# Column headers this parser recognizes, matched case-insensitively and
# stripped — not fixed column positions, since a real customer's workbook
# (verified against the one this feature was built from) has trailing
# spaces on several headers ("Transaction Type ", "Class ") and there's no
# reason to assume every future export keeps columns in exactly the same
# order.
_HEADER_ALIASES = {
    "posting_date": {"posting date"},
    "account": {"account"},
    "transaction_type": {"transaction type"},
    "class": {"class"},
    "new_class": {"new class"},
    "debit": {"debit (sar)", "debit"},
    "credit": {"credit (sar)", "credit"},
    "voucher_type": {"voucher type"},
    "voucher_no": {"voucher no", "voucher no."},
    "against_account": {"against account"},
    "project": {"project"},
    "cost_center": {"cost center", "cost centre"},
    "remarks": {"remarks"},
}


def _normalize_header(h) -> str:
    return str(h or "").strip().lower()


def _find_header_row(ws, max_scan_rows: int = 10) -> tuple[int, dict[str, int]] | None:
    """Scans the first `max_scan_rows` rows for one that contains BOTH a
    'voucher no' and a 'new class' column — the two fields this import
    cannot function without — rather than assuming a fixed row number.
    Returns (row_number, {field_key: column_index}) or None if not found."""
    for row_idx in range(1, max_scan_rows + 1):
        col_map: dict[str, int] = {}
        for col_idx in range(1, ws.max_column + 1):
            header = _normalize_header(ws.cell(row=row_idx, column=col_idx).value)
            if not header:
                continue
            for field, aliases in _HEADER_ALIASES.items():
                if header in aliases and field not in col_map:
                    col_map[field] = col_idx
        if "voucher_no" in col_map and "new_class" in col_map:
            return row_idx, col_map
    return None


def parse_classified_history_sheet(file_bytes: bytes, sheet_name: str | None = None) -> dict:
    """Returns {rows, header_row, columns_found, sheet_used, warnings}.
    `rows`: [{new_class, voucher_type, voucher_no, remarks, debit, credit,
    against_account, cost_center, project}, ...] — only rows with both a
    voucher_no and a new_class are kept; anything else is silently not a
    classified transaction row (could be a blank line, a totals row, etc.)
    and isn't reported as an error."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    warnings: list[str] = []

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_used = sheet_name
    else:
        ws = wb.worksheets[0]
        sheet_used = ws.title
        if sheet_name:
            warnings.append(f"Sheet '{sheet_name}' not found — used '{sheet_used}' instead.")

    found = _find_header_row(ws)
    if not found:
        return {"rows": [], "header_row": None, "columns_found": {}, "sheet_used": sheet_used,
                "warnings": warnings + ["Could not find a header row with both "
                                       "'Voucher No' and 'New Class' columns."]}
    header_row, col_map = found

    missing = [f for f in ("remarks", "debit", "credit", "voucher_type") if f not in col_map]
    if missing:
        warnings.append(f"Columns not found, treated as blank: {', '.join(missing)}.")

    def _cell(row_idx, field):
        col = col_map.get(field)
        return ws.cell(row=row_idx, column=col).value if col else None

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        new_class = _cell(r, "new_class")
        voucher_no = _cell(r, "voucher_no")
        if not new_class or not voucher_no:
            continue
        rows.append({
            "new_class": str(new_class).strip(),
            "voucher_type": str(_cell(r, "voucher_type") or "").strip(),
            "voucher_no": str(voucher_no).strip(),
            "remarks": str(_cell(r, "remarks") or "").strip(),
            "debit": _cell(r, "debit") or 0,
            "credit": _cell(r, "credit") or 0,
            "against_account": str(_cell(r, "against_account") or "").strip(),
            "cost_center": str(_cell(r, "cost_center") or "").strip(),
            "project": str(_cell(r, "project") or "").strip(),
        })

    return {"rows": rows, "header_row": header_row, "columns_found": col_map,
            "sheet_used": sheet_used, "warnings": warnings}


def match_lines_to_rows(rows: list[dict], available_lines: list[dict]) -> dict:
    """Pure. available_lines: [{name, label}] — the app's real Insight Cash
    Flow Line records. Matches each row's new_class (normalized: stripped,
    lowercased) against a line's label the same way. A "New Class" label
    with no matching Line is NOT silently dropped — it's reported by name
    and count, so the user knows exactly which categories need a Line
    created before those rows can be imported, the same "never silently
    absorbed" discipline as the reconciliation residual."""
    label_to_name = {(l["label"] or "").strip().lower(): l["name"] for l in available_lines}

    matched = []
    unmatched_counts: dict[str, int] = {}
    for row in rows:
        key = row["new_class"].strip().lower()
        line_name = label_to_name.get(key)
        if line_name:
            matched.append({**row, "line": line_name})
        else:
            unmatched_counts[row["new_class"]] = unmatched_counts.get(row["new_class"], 0) + 1

    return {
        "matched": matched,
        "unmatched_labels": dict(sorted(unmatched_counts.items(), key=lambda kv: -kv[1])),
        "matched_count": len(matched),
        "unmatched_count": sum(unmatched_counts.values()),
    }
