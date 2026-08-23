from __future__ import annotations

import io
import re
from typing import Any

import frappe
import openpyxl

ACC_LINE_RE = re.compile(r"^\s*(\d+)\s*-\s*(.+?)(?:\s*-\s*([A-Za-z]+))?\s*$")


def parse_account_string(s: str) -> dict | None:
    """Parse 'CODE - Name English Name Arabic - ENTITY' into a dict.

    Returns {code, name, entity} or None if it doesn't look like an account line.
    """
    if not s:
        return None
    s = str(s).strip()
    m = ACC_LINE_RE.match(s)
    if not m:
        return None
    code = m.group(1).strip()
    name = (m.group(2) or "").strip()
    entity = (m.group(3) or "").strip() or None
    return {"code": code, "name": name, "entity": entity, "raw": s}


def read_map_sheet(
    file_bytes: bytes,
    sheet_name: str = "MAP",
    *,
    account_col: int = 2,
    flag_col: int = 3,
    header_rows: int = 4,
) -> dict[str, Any]:
    """Parse the MAP sheet from an .xlsx workbook.

    The IRSAA template has these defaults: data starts at row 5, column B (idx 2) is the
    account string, column C (idx 3) is the flag.

    Returns:
        {
          "rows": [{code, name, entity, flag, raw}, ...],
          "flags": {flag: count, ...},
          "sheet_used": str,
          "warnings": [str, ...],
        }
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    warnings: list[str] = []

    if sheet_name not in wb.sheetnames:
        candidates = [s for s in wb.sheetnames if s.lower() in {"map", "mapping", "accounts"}]
        if not candidates:
            raise frappe.ValidationError(
                f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}"
            )
        sheet_name = candidates[0]
        warnings.append(f"Sheet '{sheet_name}' used as the map sheet.")

    ws = wb[sheet_name]
    rows: list[dict] = []
    flags: dict[str, int] = {}

    for row_idx, row in enumerate(ws.iter_rows(min_row=header_rows + 1, values_only=True), start=header_rows + 1):
        if row is None:
            continue
        acc_val = row[account_col - 1] if len(row) >= account_col else None
        flag_val = row[flag_col - 1] if len(row) >= flag_col else None
        parsed = parse_account_string(acc_val) if acc_val else None
        if not parsed:
            continue
        flag = (flag_val or "").strip() if flag_val else ""
        rows.append({**parsed, "flag": flag, "row_idx": row_idx})
        if flag:
            flags[flag] = flags.get(flag, 0) + 1

    if not rows:
        warnings.append("No account rows detected. Check header_rows / column indices.")

    return {"rows": rows, "flags": flags, "sheet_used": sheet_name, "warnings": warnings}


def resolve_account_to_frappe(code: str, name: str, company: str | None = None) -> str | None:
    """Find an ERPNext Account doc that matches the code (or falls back to the name).

    Strategy: account_number == code, else account_name LIKE name.
    """
    filters: dict[str, Any] = {}
    if company:
        filters["company"] = company

    if code:
        candidates = frappe.get_all(
            "Account",
            filters={**filters, "account_number": code},
            fields=["name"],
            limit_page_length=1,
        )
        if candidates:
            return candidates[0]["name"]

    if name:
        candidates = frappe.get_all(
            "Account",
            filters={**filters, "account_name": ["like", f"%{name[:40]}%"]},
            fields=["name"],
            limit_page_length=1,
        )
        if candidates:
            return candidates[0]["name"]
    return None


def apply_map_to_report(
    *,
    report: str,
    parsed: dict[str, Any],
    replace: bool = True,
    company: str | None = None,
) -> dict[str, Any]:
    """Write Account Flag Mapping rows from a parsed MAP sheet.

    If `replace`, deletes existing mappings on this report first. Otherwise,
    upserts: existing rows get their flag updated, new rows get inserted.
    Either way, this function is idempotent — running the same import twice
    produces the same end state, no DuplicateEntryError.

    Returns {created, updated, skipped_no_flag, skipped_no_match, warnings}.
    """
    # Normalize report to the DocType name (callers sometimes pass the slug).
    report_name = report
    if not frappe.db.exists("Insight Report Definition", report_name):
        # Try resolving by slug.
        by_slug = frappe.db.get_value("Insight Report Definition", {"slug": report_name}, "name")
        if by_slug:
            report_name = by_slug
        else:
            frappe.throw(f"Report '{report}' not found (neither by name nor by slug).")

    if replace:
        frappe.db.delete("Account Flag Mapping", {"report": report_name})
        frappe.db.commit()  # ensure the delete is visible before we start inserting

    created = 0
    updated = 0
    skipped_no_flag = 0
    skipped_no_match = 0
    warnings: list[str] = list(parsed.get("warnings", []))

    # Pre-load the report's existing mappings (by account_name) so we can upsert
    # instead of relying on per-row exception handling.
    existing_by_account: dict[str, str] = {
        m["account"]: m["name"]
        for m in frappe.get_all(
            "Account Flag Mapping",
            filters={"report": report_name},
            fields=["name", "account"],
            limit_page_length=0,
        )
    }

    # Track which accounts we've seen in THIS import so a duplicate row in the
    # same sheet doesn't crash the loop.
    seen_in_this_run: set[str] = set()

    for row in parsed["rows"]:
        flag = (row.get("flag") or "").strip()
        if not flag:
            skipped_no_flag += 1
            continue
        account_name = resolve_account_to_frappe(
            row.get("code") or "",
            row.get("name") or "",
            company=company,
        )
        if not account_name:
            skipped_no_match += 1
            warnings.append(
                f"No Account found for code={row.get('code')} name='{(row.get('name') or '')[:40]}'."
            )
            continue

        if account_name in seen_in_this_run:
            warnings.append(
                f"Duplicate row in MAP sheet for account '{account_name}' — kept the first."
            )
            continue
        seen_in_this_run.add(account_name)

        if account_name in existing_by_account:
            # Update in place rather than insert.
            doc = frappe.get_doc("Account Flag Mapping", existing_by_account[account_name])
            if doc.flag != flag or doc.source != "map-sheet":
                doc.flag = flag
                doc.source = "map-sheet"
                doc.auto_suggested = 0
                doc.save(ignore_permissions=True)
            updated += 1
        else:
            doc = frappe.new_doc("Account Flag Mapping")
            doc.report = report_name
            doc.account = account_name
            doc.flag = flag
            doc.source = "map-sheet"
            try:
                doc.insert(ignore_permissions=True)
                # Update our local index so subsequent rows for the same account
                # take the update path instead of trying to insert again.
                existing_by_account[account_name] = doc.name
                created += 1
            except frappe.DuplicateEntryError:
                # A race or stale index — fall back to update-by-lookup.
                fallback = frappe.db.get_value(
                    "Account Flag Mapping",
                    {"report": report_name, "account": account_name},
                    "name",
                )
                if fallback:
                    existing = frappe.get_doc("Account Flag Mapping", fallback)
                    existing.flag = flag
                    existing.source = "map-sheet"
                    existing.save(ignore_permissions=True)
                    existing_by_account[account_name] = fallback
                    updated += 1
                else:
                    # Truly unexpected — surface as warning but don't crash.
                    warnings.append(
                        f"Could not upsert mapping for account '{account_name}' (duplicate detected but no existing row found)."
                    )

    frappe.db.commit()
    return {
        "created": created,
        "updated": updated,
        "skipped_no_flag": skipped_no_flag,
        "skipped_no_match": skipped_no_match,
        "warnings": warnings,
    }
