from __future__ import annotations

import io
import re
from typing import Any

import frappe
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

CELL_REF_RE = re.compile(r"([A-Z]+)(\d+)")
EXCEL_FN_RE = re.compile(r"\b(SUM|SUMIF|SUMIFS|IF|ROUND|ABS|MIN|MAX|AVERAGE)\s*\(", re.IGNORECASE)


def import_report_structure(
    file_bytes: bytes,
    *,
    sheet_name: str = "P&L",
    label_col: int = 2,
    data_col_start: int = 4,
    flags_in_map_sheet: list[str] | None = None,
) -> dict[str, Any]:
    """Read a workbook's report sheet and infer a row tree.

    Heuristics:
      - column B (idx 2) holds the row label.
      - rows whose label matches a known P&L Classification (from the MAP sheet flags)
        become 'source' rows.
      - rows whose first data cell (col D) holds an Excel formula become 'formula' rows;
        the formula gets translated from cell refs to row-key refs using the labels
        in the same column.
      - rows whose label is non-empty but has neither flag mapping nor a formula become
        'section' headers.
      - blank rows are skipped.

    Returns:
        {
          "rows": [{key, kind, label, formula?, flag?, source_row_idx}, ...],
          "warnings": [...],
          "sheet_used": str,
        }
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False, read_only=False)

    if sheet_name not in wb.sheetnames:
        candidates = [s for s in wb.sheetnames if "p&l" in s.lower() or "pl" in s.lower() or "income" in s.lower()]
        if not candidates:
            raise frappe.ValidationError(
                f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}"
            )
        sheet_name = candidates[0]

    ws = wb[sheet_name]
    flags = set(flags_in_map_sheet or [])
    warnings: list[str] = []

    row_specs: list[dict] = []
    label_to_key: dict[str, str] = {}

    for row_idx in range(1, ws.max_row + 1):
        label_cell = ws.cell(row=row_idx, column=label_col)
        label = label_cell.value
        if not isinstance(label, str):
            continue
        label = label.strip()
        if not label or label.startswith("="):
            continue

        first_data_cell = ws.cell(row=row_idx, column=data_col_start)
        first_data_val = first_data_cell.value
        is_formula = isinstance(first_data_val, str) and first_data_val.startswith("=")

        kind: str
        formula_for_row: str | None = None

        if is_formula:
            kind = "formula"
            formula_for_row = first_data_val[1:]
        elif label in flags:
            kind = "source"
        else:
            looks_like_section = (label.endswith(":") or label.istitle() or _is_header_label(label))
            kind = "section" if looks_like_section else "source"
            if kind == "source":
                warnings.append(
                    f"Row {row_idx} '{label}' was treated as source but has no MAP flag yet."
                )

        key = _slug(label, used=label_to_key.values())
        label_to_key[label] = key
        spec: dict[str, Any] = {
            "key": key,
            "kind": kind,
            "label": label,
            "source_row_idx": row_idx,
        }
        if kind == "source":
            spec["flag"] = label
            spec["accounts"] = []
        if formula_for_row is not None:
            spec["raw_excel_formula"] = formula_for_row
        row_specs.append(spec)

    label_by_row_idx = {s["source_row_idx"]: s["label"] for s in row_specs}
    key_by_row_idx = {s["source_row_idx"]: s["key"] for s in row_specs}

    for spec in row_specs:
        if spec["kind"] != "formula":
            continue
        excel_formula = spec.get("raw_excel_formula") or ""
        translated, unresolved = translate_formula(
            excel_formula,
            data_col_start=data_col_start,
            key_by_row_idx=key_by_row_idx,
        )
        spec["formula"] = translated
        if unresolved:
            warnings.append(
                f"Formula row '{spec['label']}' has unresolved refs: {unresolved}"
            )

    return {
        "rows": row_specs,
        "warnings": warnings,
        "sheet_used": sheet_name,
    }


def translate_formula(
    excel_formula: str,
    *,
    data_col_start: int,
    key_by_row_idx: dict[int, str],
) -> tuple[str, list[str]]:
    """Convert an Excel formula like '=D8-D10-D12' (or '=SUM(D17:D25)') into a row-key formula.

    Strategy: cell refs in the data column are replaced with the row key for that row index.
    SUM(rangeStartRow:rangeEndRow) is expanded to a sum of every row key in the range.
    """
    unresolved: list[str] = []
    expanded = _expand_sum_ranges(excel_formula, data_col_start=data_col_start, key_by_row_idx=key_by_row_idx)

    def replace_cell(match: re.Match) -> str:
        col_letter = match.group(1)
        row_idx = int(match.group(2))
        col_idx = column_index_from_string(col_letter)
        key = key_by_row_idx.get(row_idx)
        if not key:
            unresolved.append(f"{col_letter}{row_idx}")
            return "0"
        return key

    translated = CELL_REF_RE.sub(replace_cell, expanded)
    translated = re.sub(r"\bSUM\s*\(", "SUM(", translated, flags=re.IGNORECASE)
    translated = re.sub(r"^\s*\+", "", translated)
    return translated.strip(), unresolved


def _expand_sum_ranges(formula: str, *, data_col_start: int, key_by_row_idx: dict[int, str]) -> str:
    range_re = re.compile(r"(?P<col1>[A-Z]+)(?P<r1>\d+):(?P<col2>[A-Z]+)(?P<r2>\d+)")

    def expand(m: re.Match) -> str:
        r1 = int(m.group("r1"))
        r2 = int(m.group("r2"))
        if r1 > r2:
            r1, r2 = r2, r1
        keys = []
        for ri in range(r1, r2 + 1):
            k = key_by_row_idx.get(ri)
            if k:
                keys.append(k)
        if not keys:
            return "0"
        return "(" + " + ".join(keys) + ")"

    return range_re.sub(expand, formula)


def _is_header_label(label: str) -> bool:
    if any(w in label.lower() for w in ["total", "subtotal", "gross", "net ", "ebitda", "pbt", "pat"]):
        return False
    if label.startswith(("Sales ", "Cost ", "Revenue", "Other ")):
        return False
    words = label.split()
    if len(words) <= 4 and not any(c.isdigit() for c in label):
        return True
    return False


def _slug(label: str, used) -> str:
    used_set = set(used)
    base = re.sub(r"[^a-z0-9]+", "_", label.lower()).strip("_")
    if not base:
        base = "row"
    candidate = base
    n = 2
    while candidate in used_set:
        candidate = f"{base}_{n}"
        n += 1
    return candidate
