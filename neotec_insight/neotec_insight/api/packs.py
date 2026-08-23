# Copyright (c) 2026, Neotec Integrated Solution
# Export Pack engine (v2.28.0) — client-specific audit workbooks WITHOUT
# client-specific code. A pack is CONFIGURATION: an ordered list of SHEET
# COMPONENTS from the catalog below, each with schema-driven options. The
# frontend designer renders the option UI from the schema, so adding a new
# sheet type here automatically appears in the designer — the next client's
# "different requirement" is a new pack, not a build.
#
# Catalog: vat_return (16-box summary) · sales_register (invoice/item level,
# customer VAT no.) · purchase_register (incl. non-invoice sources) ·
# gl_ledger (any accounts, or the strict-tagged VAT accounts, with running
# balance). Bilingual EN/AR headers.
from __future__ import annotations

import io
import json

import frappe
from frappe import _
from frappe.utils import flt

from .health import _default_company
from .vat import (STANDARD_RATE, _apply_adjustments, _non_invoice_vouchers,
                  _vat_accounts, vat_return)

# ── Sheet-component catalog (schema drives the designer UI) ─────────────────
SHEET_TYPES = [
    {
        "type": "vat_return",
        "label": "VAT Return (16 boxes)",
        "options": [],
        "columns": [],
    },
    {
        "type": "sales_register",
        "label": "Sales VAT Register",
        "options": [
            {"key": "item_level", "label": "One row per item line", "kind": "check", "default": 1},
            {"key": "include_returns", "label": "Include credit notes (negative rows)", "kind": "check", "default": 1},
        ],
        "columns": [
            ("idx", "#", "م"), ("party", "Customer Name", "اسم العميل"),
            ("tax_id", "VAT No.", "الرقم الضريبي"), ("voucher_no", "Invoice No.", "رقم الفاتورة"),
            ("posting_date", "Issue Date", "تاريخ الإصدار"), ("item", "Product / Description", "المنتج / الوصف"),
            ("qty", "Qty", "الكمية"), ("rate", "Unit Price", "سعر الوحدة"),
            ("net", "Amount", "المبلغ"), ("tax_rate", "Tax Rate", "نسبة الضريبة"),
            ("vat", "VAT Amount", "مبلغ الضريبة"), ("grand", "Grand Total", "الإجمالي"),
            ("outstanding", "Outstanding", "المتبقي"),
        ],
    },
    {
        "type": "purchase_register",
        "label": "Purchases VAT Register",
        "options": [
            {"key": "item_level", "label": "One row per item line", "kind": "check", "default": 1},
            {"key": "include_non_invoice", "label": "Include non-invoice documents (Expenses Entry, Payment Entry…)", "kind": "check", "default": 1},
        ],
        "columns": [
            ("idx", "#", "م"), ("party", "Vendor Name", "اسم المورد"),
            ("tax_id", "VAT No.", "الرقم الضريبي"), ("bill_no", "Invoice No.", "رقم الفاتورة"),
            ("bill_date", "Invoice Date", "تاريخ الفاتورة"), ("posting_date", "Posting Date", "تاريخ القيد"),
            ("item", "Product / Description", "المنتج / الوصف"),
            ("qty", "Qty", "الكمية"), ("rate", "Unit Price", "سعر الوحدة"),
            ("net", "Amount", "المبلغ"), ("tax_rate", "Tax Rate", "نسبة الضريبة"),
            ("vat", "VAT Amount", "مبلغ الضريبة"), ("source", "Source", "المصدر"),
        ],
    },
    {
        "type": "gst_sales_register",
        "label": "GST Sales Register (B2B)",
        "options": [],
        "columns": [
            ("idx", "#", "#"), ("party", "Customer", "Customer"),
            ("gstin", "GSTIN", "GSTIN"), ("voucher_no", "Invoice No.", "Invoice No."),
            ("posting_date", "Invoice Date", "Invoice Date"),
            ("net", "Taxable Value", "Taxable Value"),
            ("cgst", "CGST", "CGST"), ("sgst", "SGST/UTGST", "SGST/UTGST"),
            ("igst", "IGST", "IGST"), ("cess", "Cess", "Cess"),
            ("grand", "Invoice Total", "Invoice Total"),
        ],
    },
    {
        "type": "gst_purchase_register",
        "label": "GST Purchase Register (ITC)",
        "options": [],
        "columns": [
            ("idx", "#", "#"), ("party", "Supplier", "Supplier"),
            ("gstin", "GSTIN", "GSTIN"), ("bill_no", "Bill No.", "Bill No."),
            ("posting_date", "Posting Date", "Posting Date"),
            ("net", "Taxable Value", "Taxable Value"),
            ("cgst", "CGST", "CGST"), ("sgst", "SGST/UTGST", "SGST/UTGST"),
            ("igst", "IGST", "IGST"), ("cess", "Cess", "Cess"),
            ("grand", "Bill Total", "Bill Total"),
        ],
    },
    {
        "type": "gl_ledger",
        "label": "GL Ledger (any accounts)",
        "options": [
            {"key": "accounts_mode", "label": "Accounts", "kind": "select", "default": "output_vat",
             "choices": [["output_vat", "Output VAT accounts (tagged/detected)"],
                          ["input_vat", "Input VAT accounts (tagged/detected)"],
                          ["custom", "Custom account list"]]},
            {"key": "accounts", "label": "Custom accounts (comma-separated)", "kind": "text", "default": ""},
            {"key": "include_opening", "label": "Print the opening balance row", "kind": "check", "default": 1},
        ],
        "columns": [
            ("posting_date", "Posting Date", "تاريخ القيد"), ("account", "Account", "الحساب"),
            ("debit", "Debit", "مدين"), ("credit", "Credit", "دائن"),
            ("balance", "Balance", "الرصيد"), ("voucher_type", "Voucher Type", "نوع المستند"),
            ("voucher_no", "Voucher No", "رقم المستند"), ("against", "Against Account", "الحساب المقابل"),
            ("party", "Party", "الطرف"), ("remarks", "Remarks", "ملاحظات"),
        ],
    },
    {
        "type": "gtpl_register",
        "label": "Government VAT Deferral Register (GTPL)",
        "options": [
            {"key": "show_attention", "label": "Show rows needing a decision (part-paid, ungrouped)", "kind": "check", "default": 1},
            {"key": "show_already_counted", "label": "Show supplies already declared in earlier returns", "kind": "check", "default": 0},
        ],
        "columns": [
            ("idx", "#", "م"), ("voucher_no", "Invoice No.", "رقم الفاتورة"),
            ("posting_date", "Issue Date", "تاريخ الإصدار"),
            ("customer", "Customer", "العميل"),
            ("release_date", "Tax Due Date", "تاريخ استحقاق الضريبة"),
            ("net", "Amount", "المبلغ"), ("vat", "VAT Amount", "مبلغ الضريبة"),
            ("why", "Basis", "الأساس"),
        ],
    },
]
_TYPES = {t["type"]: t for t in SHEET_TYPES}


def _require_read() -> None:
    """Refuse a financial read from a user with no ledger access.

    `@frappe.whitelist()` requires a login, not a role, so without this these
    endpoints were callable over `/api/method/...` by any authenticated user,
    including portal users with no business seeing the ledger. Reading GL Entry
    is the right test: ERPNext already restricts it to the accounts roles, so
    this inherits the site's own configuration rather than inventing a second
    permission model.
    """
    if not frappe.has_permission("GL Entry", "read"):
        frappe.throw(
            _("You are not permitted to view financial data."),
            frappe.PermissionError,
        )


@frappe.whitelist()
def list_sheet_types():
    _require_read()
    return SHEET_TYPES


@frappe.whitelist()
def list_packs():
    _require_read()
    return frappe.get_all("Insight Export Pack",
                          fields=["slug", "title", "description", "modified"],
                          order_by="modified desc", limit_page_length=100)


@frappe.whitelist()
def load_pack(slug=None):
    _require_read()
    if not slug or not frappe.db.exists("Insight Export Pack", slug):
        frappe.throw(_("Pack not found."))
    doc = frappe.get_doc("Insight Export Pack", slug)
    return {"slug": doc.slug, "title": doc.title, "description": doc.description,
            "config": json.loads(doc.config_json or "{}")}


@frappe.whitelist()
def save_pack(pack=None):
    if not frappe.has_permission("Insight Export Pack", "write"):
        frappe.throw(_("Not permitted."))
    if isinstance(pack, str):
        pack = json.loads(pack or "{}")
    title = (pack.get("title") or "").strip()
    if not title:
        frappe.throw(_("Give the pack a title."))
    cfg = pack.get("config") or {}
    sheets = [s for s in (cfg.get("sheets") or []) if s.get("type") in _TYPES][:12]
    if not sheets:
        frappe.throw(_("A pack needs at least one sheet."))
    cfg = {"sheets": sheets, "language": cfg.get("language") or "both"}
    slug = pack.get("slug") or frappe.scrub(title)
    if frappe.db.exists("Insight Export Pack", slug):
        doc = frappe.get_doc("Insight Export Pack", slug)
    else:
        doc = frappe.new_doc("Insight Export Pack")
        doc.slug = slug
    doc.title = title
    doc.description = pack.get("description") or ""
    doc.config_json = json.dumps(cfg)
    doc.save(ignore_permissions=True)
    frappe.db.commit()
    return {"slug": doc.slug}


@frappe.whitelist()
def delete_pack(slug=None):
    if not frappe.has_permission("Insight Export Pack", "delete"):
        frappe.throw(_("Not permitted."))
    if slug and frappe.db.exists("Insight Export Pack", slug):
        frappe.delete_doc("Insight Export Pack", slug)
        frappe.db.commit()
    return {"ok": True}


def seed_default_export_packs() -> None:
    """Idempotent seed (called from install after_migrate): the 'VAT Breakdown
    (Quarterly)' pack mirroring the live client's manual workbook — return
    summary, item-level sales & purchase registers, and both VAT-account GL
    ledgers. Never overwrites a pack the user has edited."""
    if frappe.db.exists("Insight Export Pack", "vat_breakdown_quarterly"):
        return
    cfg = {"language": "both", "sheets": [
        {"type": "vat_return", "title": "VAT Return", "options": {}},
        {"type": "sales_register", "title": "Sales VAT", "options": {"item_level": 1, "include_returns": 1}},
        {"type": "purchase_register", "title": "Purchases VAT", "options": {"item_level": 1, "include_non_invoice": 1}},
        {"type": "gl_ledger", "title": "GL - Output VAT", "options": {"accounts_mode": "output_vat"}},
        {"type": "gl_ledger", "title": "GL - Input VAT", "options": {"accounts_mode": "input_vat"}},
    ]}
    frappe.get_doc({
        "doctype": "Insight Export Pack", "slug": "vat_breakdown_quarterly",
        "title": "VAT Breakdown (Quarterly)",
        "description": "Audit pack: 16-box return + item-level sales/purchase VAT registers with party VAT numbers + Output/Input VAT GL ledgers.",
        "config_json": json.dumps(cfg),
    }).insert(ignore_permissions=True)


# ── Data builders ────────────────────────────────────────────────────────────

def _pct(v):
    return flt(v, 4)


def _sales_rows(company, from_date, to_date, opts):
    filters = {"company": company, "docstatus": 1,
               "posting_date": ["between", [from_date, to_date]]}
    if not opts.get("include_returns", 1):
        filters["is_return"] = 0
    _fields = ["name", "customer", "customer_name", "tax_id",
               "posting_date", "base_net_total",
               "base_total_taxes_and_charges", "base_grand_total",
               "outstanding_amount", "is_return"]
    invs = frappe.get_all("Sales Invoice", filters=filters, fields=_fields,
                          order_by="posting_date asc, name asc", limit_page_length=0)
    invs, excluded = _apply_adjustments(invs, "Sales Invoice", company, from_date, to_date, _fields)
    invs = invs + excluded  # excluded stay visible (red) but are not totalled
    rows = []
    if opts.get("item_level", 1):
        by_parent = {}
        if invs:
            items = frappe.get_all("Sales Invoice Item",
                                   filters={"parent": ["in", [i["name"] for i in invs]]},
                                   fields=["parent", "item_name", "qty", "base_rate", "base_net_amount"],
                                   parent_doctype="Sales Invoice", order_by="parent, idx",
                                   limit_page_length=0)
            for it in items:
                by_parent.setdefault(it["parent"], []).append(it)
        for inv in invs:
            eff = (inv["base_total_taxes_and_charges"] or 0) / inv["base_net_total"] if inv["base_net_total"] else 0
            for it in by_parent.get(inv["name"], [{"item_name": "", "qty": None, "base_rate": None, "base_net_amount": inv["base_net_total"]}]):
                net = flt(it.get("base_net_amount"), 2)
                rows.append({"_adj": inv.get("_adj"), "_adj_reason": inv.get("_adj_reason"),
                             "party": inv["customer_name"] or inv["customer"],
                             "tax_id": inv.get("tax_id") or "",
                             "voucher_no": inv["name"], "posting_date": inv["posting_date"],
                             "item": it.get("item_name") or "", "qty": it.get("qty"),
                             "rate": flt(it.get("base_rate"), 2) if it.get("base_rate") is not None else None,
                             "net": net, "tax_rate": _pct(eff),
                             "vat": flt(net * eff, 2),
                             "grand": flt(inv["base_grand_total"], 2),
                             "outstanding": flt(inv["outstanding_amount"], 2)})
        return rows
    for inv in invs:
        eff = (inv["base_total_taxes_and_charges"] or 0) / inv["base_net_total"] if inv["base_net_total"] else 0
        rows.append({"_adj": inv.get("_adj"), "_adj_reason": inv.get("_adj_reason"),
                     "party": inv["customer_name"] or inv["customer"], "tax_id": inv.get("tax_id") or "",
                     "voucher_no": inv["name"], "posting_date": inv["posting_date"], "item": "",
                     "qty": None, "rate": None, "net": flt(inv["base_net_total"], 2),
                     "tax_rate": _pct(eff), "vat": flt(inv["base_total_taxes_and_charges"], 2),
                     "grand": flt(inv["base_grand_total"], 2),
                     "outstanding": flt(inv["outstanding_amount"], 2)})
    return rows


def _purchase_rows(company, from_date, to_date, opts):
    _pfields = ["name", "supplier", "supplier_name", "tax_id",
                "bill_no", "bill_date", "posting_date", "base_net_total",
                "base_total_taxes_and_charges"]
    invs = frappe.get_all("Purchase Invoice",
                          filters={"company": company, "docstatus": 1,
                                   "posting_date": ["between", [from_date, to_date]]},
                          fields=_pfields,
                          order_by="posting_date asc, name asc", limit_page_length=0)
    invs, p_excluded = _apply_adjustments(invs, "Purchase Invoice", company, from_date, to_date, _pfields)
    invs = invs + p_excluded
    rows = []
    if opts.get("item_level", 1) and invs:
        items = frappe.get_all("Purchase Invoice Item",
                               filters={"parent": ["in", [i["name"] for i in invs]]},
                               fields=["parent", "item_name", "qty", "base_rate", "base_net_amount"],
                               parent_doctype="Purchase Invoice", order_by="parent, idx",
                               limit_page_length=0)
        by_parent = {}
        for it in items:
            by_parent.setdefault(it["parent"], []).append(it)
        for inv in invs:
            eff = (inv["base_total_taxes_and_charges"] or 0) / inv["base_net_total"] if inv["base_net_total"] else 0
            for it in by_parent.get(inv["name"], [{"item_name": "", "qty": None, "base_rate": None, "base_net_amount": inv["base_net_total"]}]):
                net = flt(it.get("base_net_amount"), 2)
                rows.append({"_adj": inv.get("_adj"), "_adj_reason": inv.get("_adj_reason"),
                             "party": inv["supplier_name"] or inv["supplier"],
                             "tax_id": inv.get("tax_id") or "",
                             "bill_no": inv.get("bill_no") or inv["name"],
                             "bill_date": inv.get("bill_date") or inv["posting_date"],
                             "posting_date": inv["posting_date"],
                             "item": it.get("item_name") or "", "qty": it.get("qty"),
                             "rate": flt(it.get("base_rate"), 2) if it.get("base_rate") is not None else None,
                             "net": net, "tax_rate": _pct(eff), "vat": flt(net * eff, 2),
                             "source": "Purchase Invoice"})
    else:
        for inv in invs:
            eff = (inv["base_total_taxes_and_charges"] or 0) / inv["base_net_total"] if inv["base_net_total"] else 0
            rows.append({"_adj": inv.get("_adj"), "_adj_reason": inv.get("_adj_reason"),
                         "party": inv["supplier_name"] or inv["supplier"], "tax_id": inv.get("tax_id") or "",
                         "bill_no": inv.get("bill_no") or inv["name"],
                         "bill_date": inv.get("bill_date") or inv["posting_date"],
                         "posting_date": inv["posting_date"], "item": "", "qty": None, "rate": None,
                         "net": flt(inv["base_net_total"], 2), "tax_rate": _pct(eff),
                         "vat": flt(inv["base_total_taxes_and_charges"], 2),
                         "source": "Purchase Invoice"})
    if opts.get("include_non_invoice", 1):
        out_accts, in_accts, clearing_accts = _vat_accounts(company)
        rate = STANDARD_RATE / 100.0
        for v in _non_invoice_vouchers(
                company, in_accts, out_accts + clearing_accts, from_date, to_date, debit_positive=True):
            rows.append({"party": v.get("supplier") or "", "tax_id": "",
                         "bill_no": v["name"], "bill_date": v["posting_date"],
                         "posting_date": v["posting_date"], "item": "", "qty": None, "rate": None,
                         "net": flt(v["base_total_taxes_and_charges"] / rate, 2),
                         "tax_rate": _pct(rate), "vat": v["base_total_taxes_and_charges"],
                         "source": v["doctype"]})
        rows.sort(key=lambda r: str(r["posting_date"]))
    return rows


def _gl_rows(company, from_date, to_date, opts):
    mode = (opts.get("accounts_mode") or "output_vat")
    if mode == "custom":
        accounts = [a.strip() for a in str(opts.get("accounts") or "").split(",") if a.strip()]
    else:
        out_accts, in_accts, _clearing_accts = _vat_accounts(company)
        accounts = out_accts if mode == "output_vat" else in_accts
    if not accounts:
        return [], []
    opening = frappe.db.sql(
        """SELECT COALESCE(SUM(debit),0) d, COALESCE(SUM(credit),0) c FROM `tabGL Entry`
           WHERE company=%(c)s AND is_cancelled=0 AND account IN %(a)s
             AND posting_date < %(f)s""",
        {"c": company, "a": tuple(accounts), "f": from_date}, as_dict=True)[0]
    rows = frappe.db.sql(
        """SELECT posting_date, account, debit, credit, voucher_type, voucher_no,
                  against, party, remarks
           FROM `tabGL Entry`
           WHERE company=%(c)s AND is_cancelled=0 AND account IN %(a)s
             AND posting_date BETWEEN %(f)s AND %(t)s
           ORDER BY posting_date, creation""",
        {"c": company, "a": tuple(accounts), "f": from_date, "t": to_date}, as_dict=True)
    bal = flt(opening["d"] - opening["c"], 2)
    out = []
    if int(opts.get("include_opening", 1)):
        out.append({"posting_date": "", "account": _("Opening Balance"),
                    "debit": flt(opening["d"], 2), "credit": flt(opening["c"], 2),
                    "balance": bal, "voucher_type": "", "voucher_no": "",
                    "against": "", "party": "", "remarks": ""})
    for r in rows:
        bal = flt(bal + flt(r["debit"]) - flt(r["credit"]), 2)
        out.append({"posting_date": r["posting_date"], "account": r["account"],
                    "debit": flt(r["debit"], 2), "credit": flt(r["credit"], 2), "balance": bal,
                    "voucher_type": r["voucher_type"] or "", "voucher_no": r["voucher_no"] or "",
                    "against": r.get("against") or "", "party": r.get("party") or "",
                    "remarks": (r.get("remarks") or "")[:120]})
    return out, accounts


def _gst_head_of(account_head):
    from .gst import _HEAD_RX
    for h, rx in _HEAD_RX.items():
        if rx.search(account_head or ""):
            return h
    return None


def _gst_register_rows(company, from_date, to_date, doctype):
    """Per-invoice CGST/SGST/IGST/Cess split read from the invoice taxes
    child table (account_head name convention)."""
    tax_table = ("Sales Taxes and Charges" if doctype == "Sales Invoice"
                 else "Purchase Taxes and Charges")
    party_f = "customer_name" if doctype == "Sales Invoice" else "supplier_name"
    fields = ["name", party_f + " as party", "tax_id", "posting_date",
              "base_net_total", "base_grand_total"]
    if doctype == "Purchase Invoice":
        fields.append("bill_no")
    invs = frappe.get_all(doctype, filters={"company": company, "docstatus": 1,
                          "posting_date": ["between", [from_date, to_date]]},
                          fields=fields, order_by="posting_date asc, name asc",
                          limit_page_length=0)
    heads_by_parent = {}
    if invs:
        taxes = frappe.get_all(tax_table,
                               filters={"parent": ["in", [i["name"] for i in invs]]},
                               fields=["parent", "account_head", "base_tax_amount_after_discount_amount as amt"],
                               parent_doctype=doctype, limit_page_length=0)
        for tx in taxes:
            h = _gst_head_of(tx["account_head"])
            if h:
                d = heads_by_parent.setdefault(tx["parent"], {})
                d[h] = flt(d.get(h, 0) + flt(tx["amt"]), 2)
    rows = []
    for inv in invs:
        h = heads_by_parent.get(inv["name"], {})
        rows.append({"party": inv["party"], "gstin": inv.get("tax_id") or "",
                     "voucher_no": inv["name"], "bill_no": inv.get("bill_no") or inv["name"],
                     "posting_date": inv["posting_date"],
                     "net": flt(inv["base_net_total"], 2),
                     "cgst": flt(h.get("cgst", 0), 2), "sgst": flt(h.get("sgst", 0), 2),
                     "igst": flt(h.get("igst", 0), 2), "cess": flt(h.get("cess", 0), 2),
                     "grand": flt(inv["base_grand_total"], 2)})
    return rows


# ── Workbook generator ───────────────────────────────────────────────────────

@frappe.whitelist()
def generate_pack(slug=None, company=None, from_date=None, to_date=None, letter_head=None):
    from datetime import datetime

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
    from openpyxl.utils import get_column_letter

    company = company or _default_company()
    if not from_date or not to_date:
        frappe.throw(_("from_date and to_date are required."))
    if not frappe.has_permission("GL Entry", "read"):
        frappe.throw(_("Not permitted."))
    pack = load_pack(slug)
    cfg = pack["config"]
    lang = cfg.get("language") or "both"

    def H(en, ar):
        return en if lang == "en" else ar if lang == "ar" else f"{en} / {ar}"

    wb = Workbook()
    wb.remove(wb.active)
    F_HEAD = PatternFill("solid", fgColor="1F4E5F")
    F_TITLE = Font(bold=True, size=14, color="1F4E5F")
    F_TOT = PatternFill("solid", fgColor="EEF7F4")
    thin = Border(bottom=Side(style="thin", color="E0E0E0"))

    def new_sheet(title):
        ws = wb.create_sheet(title=title[:31] or "Sheet")
        ws.sheet_view.showGridLines = False
        return ws

    def head_block(ws, title):
        ws.append([title]); ws["A1"].font = F_TITLE
        ws.append([f"{company} · {from_date} → {to_date}"])
        ws["A2"].font = Font(size=9, color="777777")
        ws.append([])

    def header_row(ws, labels):
        ws.append(labels)
        for c in ws[ws.max_row]:
            c.fill = F_HEAD; c.font = Font(bold=True, color="FFFFFF")

    def autow(ws):
        for j, col in enumerate(ws.columns, start=1):
            w = 10
            for c in list(col)[:60]:
                w = max(w, min(len(str(c.value or "")), 42))
            ws.column_dimensions[get_column_letter(j)].width = w + 2

    def _dcell(ws, v):
        if hasattr(v, "isoformat") or (isinstance(v, str) and len(v) == 10 and v[4:5] == "-"):
            try:
                return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
            except Exception:
                return v
        return v

    F_INC = PatternFill("solid", fgColor="DFF0D8")   # included from other period
    F_EXC = PatternFill("solid", fgColor="FDECEA")   # excluded / deferred

    def register_sheet(sheet_cfg, rows, spec):
        ws = new_sheet(sheet_cfg.get("title") or spec["label"])
        head_block(ws, sheet_cfg.get("title") or spec["label"])
        cols = spec["columns"]
        chosen = sheet_cfg.get("options", {}).get("columns") or [c[0] for c in cols]
        cols = [c for c in cols if c[0] in chosen]
        header_row(ws, [H(c[1], c[2]) for c in cols])
        num_keys = {"qty", "rate", "net", "vat", "grand", "outstanding", "debit", "credit", "balance",
                    "cgst", "sgst", "igst", "cess"}
        totals = {k: 0.0 for k in ("net", "vat", "grand", "outstanding", "debit", "credit",
                                    "cgst", "sgst", "igst", "cess")}
        recon = {"in": {"vat": 0.0, "net": 0.0, "rows": 0},
                 "out": {"vat": 0.0, "net": 0.0, "rows": 0}}
        for i, r in enumerate(rows, start=1):
            adj = r.get("_adj")
            vals = []
            for (k, _en, _ar) in cols:
                v = i if k == "idx" else r.get(k)
                if k in ("posting_date", "bill_date"):
                    v = _dcell(ws, v)
                vals.append(v)
            ws.append(vals)
            for j, (k, _en, _ar) in enumerate(cols, start=1):
                cell = ws.cell(row=ws.max_row, column=j)
                cell.border = thin
                if adj == "in":
                    cell.fill = F_INC
                elif adj == "out":
                    cell.fill = F_EXC
                    cell.font = Font(color="B3261E")
                if k in ("posting_date", "bill_date"):
                    cell.number_format = "DD-MM-YYYY"
                elif k == "tax_rate":
                    cell.number_format = "0%"
                elif k in num_keys:
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right")
                # Excluded rows are visible but never counted.
                if adj != "out" and k in totals and isinstance(r.get(k), (int, float)):
                    totals[k] += r[k]
            if adj in recon:
                recon[adj]["vat"] += flt(r.get("vat") or 0)
                recon[adj]["net"] += flt(r.get("net") or 0)
                recon[adj]["rows"] += 1
        # Totals row
        vals = []
        for (k, _en, _ar) in cols:
            vals.append(_("TOTAL") if k in ("idx", "party", "posting_date") and not vals else
                        (flt(totals[k], 2) if k in totals else None))
        ws.append(vals)
        for j, (k, _en, _ar) in enumerate(cols, start=1):
            c = ws.cell(row=ws.max_row, column=j)
            c.fill = F_TOT; c.font = Font(bold=True, color="11816F")
            if k in num_keys:
                c.number_format = "#,##0.00"; c.alignment = Alignment(horizontal="right")
        # Reconciliation block (v2.34.0) — a proper labeled mini-table whose
        # arithmetic CHAINS: invoiced-in-period (including rows later
        # deferred) − deferred + included-from-other-periods = VAT for this
        # return. Previous layout dropped numbers under unrelated register
        # columns and mixed a non-chaining figure — confusing, now fixed.
        if recon["in"]["rows"] or recon["out"]["rows"]:
            gross_net = flt(totals.get("net", 0) - recon["in"]["net"] + recon["out"]["net"], 2)
            gross_vat = flt(totals.get("vat", 0) - recon["in"]["vat"] + recon["out"]["vat"], 2)
            ws.append([])
            ws.append([H("Reconciliation", "التسوية"),
                       H("Amount", "المبلغ"), H("VAT", "الضريبة")])
            for c in ws[ws.max_row][:3]:
                c.fill = F_HEAD; c.font = Font(bold=True, color="FFFFFF")
            lines = [(H("Invoiced in period", "فواتير الفترة"), gross_net, gross_vat, False)]
            if recon["out"]["rows"]:
                lines.append((H("Less: deferred to payment period",
                                "ناقص: مؤجلة إلى فترة السداد"),
                              flt(-recon["out"]["net"], 2), flt(-recon["out"]["vat"], 2), False))
            if recon["in"]["rows"]:
                lines.append((H("Add: included from other periods (paid this period)",
                                "زائد: مضافة من فترات أخرى (مسددة خلال الفترة)"),
                              flt(recon["in"]["net"], 2), flt(recon["in"]["vat"], 2), False))
            lines.append((H("VAT for this return", "الضريبة لهذا الإقرار"),
                          flt(totals.get("net", 0), 2), flt(totals.get("vat", 0), 2), True))
            for (label, netv, vatv, last) in lines:
                ws.append([label, netv, vatv])
                for j in (1, 2, 3):
                    c = ws.cell(row=ws.max_row, column=j)
                    c.font = Font(bold=last, italic=not last,
                                  color="11816F" if last else "555555")
                    if j > 1:
                        c.number_format = "#,##0.00"
                        c.alignment = Alignment(horizontal="right")
                    if last:
                        c.fill = F_TOT
        autow(ws)

    def gtpl_sheet(sheet_cfg, spec, opts):
        """The deferral register: which government supplies this return declares,
        which it carries forward, and which need a human before it can be filed.

        Sectioned rather than flat because the three populations answer different
        questions and get summed differently — declared feeds the box Amount,
        credit notes feed its Adjustment, and deferred feeds neither while still
        needing to be visible and to reconcile."""
        from ..utils.gtpl_core import box_figures
        from .gtpl import build_plan, ledger_check

        plan = build_plan(company, from_date, to_date)
        ws = new_sheet(sheet_cfg.get("title") or spec["label"])
        head_block(ws, sheet_cfg.get("title") or spec["label"])

        rule = plan.get("rule")
        if not rule:
            ws.append([_("No active GTPL rule applies to a period ending {0}.").format(to_date)])
            ws[f"A{ws.max_row}"].font = Font(italic=True, color="B3261E")
            ws.append([_("Create an Insight GTPL Rule effective on or before that date.")])
            ws[f"A{ws.max_row}"].font = Font(size=9, color="777777")
            autow(ws)
            return
        ws.append(["{0}: {1} · {2} · {3} {4}".format(
            _("Rule"), rule["name"], rule["trigger_basis"],
            _("target box"), rule.get("target_box") or "—")])
        ws[f"A{ws.max_row}"].font = Font(size=9, color="777777")
        ws.append([])

        cols = spec["columns"]
        show_attention = int(opts.get("show_attention", 1) or 0)
        show_counted = int(opts.get("show_already_counted", 0) or 0)

        sections = [
            (H("Declared in this return", "المصرح به في هذا الإقرار"),
             {"released", "released_by_credit_note"}, F_INC, None),
            (H("Adjustments — credit notes issued this period", "التعديلات — إشعارات دائنة صادرة خلال الفترة"),
             {"credit_note"}, PatternFill("solid", fgColor="FFF6E5"), None),
            (H("Carried forward — not yet due", "مرحّل — لم تستحق بعد"),
             {"deferred", "still_deferred"}, F_EXC, "B3261E"),
        ]
        if show_attention:
            sections.append((H("Needs a decision before filing", "تحتاج قراراً قبل التقديم"),
                             {"partial", "scope_unknown"},
                             PatternFill("solid", fgColor="F0F0F0"), "555555"))
        if show_counted:
            sections.append((H("Already declared in earlier returns", "سبق التصريح بها في إقرارات سابقة"),
                             {"already_counted", "in_period"},
                             PatternFill("solid", fgColor="F7F7F7"), "777777"))

        subtotals = {}
        for (label, states, fill, colour) in sections:
            rows = [r for r in plan["rows"] if r["state"] in states]
            ws.append([label])
            c = ws.cell(row=ws.max_row, column=1)
            c.font = Font(bold=True, color="1F4E5F"); c.fill = fill
            if not rows:
                ws.append([_("None.")])
                ws[f"A{ws.max_row}"].font = Font(italic=True, size=9, color="777777")
                subtotals[label] = (0.0, 0.0)
                ws.append([])
                continue
            header_row(ws, [H(c1, c2) for (_k, c1, c2) in cols])
            net = vat = 0.0
            for i, r in enumerate(rows, start=1):
                ws.append([i if k == "idx" else
                           (_dcell(ws, r.get(k)) if k in ("posting_date", "release_date") else r.get(k))
                           for (k, _e, _a) in cols])
                for j, (k, _e, _a) in enumerate(cols, start=1):
                    cell = ws.cell(row=ws.max_row, column=j)
                    cell.border = thin
                    cell.fill = fill
                    if colour:
                        cell.font = Font(color=colour)
                    if k in ("posting_date", "release_date"):
                        cell.number_format = "DD-MM-YYYY"
                    elif k in ("net", "vat"):
                        cell.number_format = "#,##0.00"
                        cell.alignment = Alignment(horizontal="right")
                net += flt(r.get("net"), 2); vat += flt(r.get("vat"), 2)
            ws.append([None] * (len(cols) - 3) + [H("Subtotal", "المجموع الفرعي"),
                                                  flt(net, 2), flt(vat, 2)])
            for cc in ws[ws.max_row]:
                cc.font = Font(bold=True); cc.fill = F_TOT
            for j in (len(cols) - 1, len(cols)):
                ws.cell(row=ws.max_row, column=j).number_format = "#,##0.00"
            subtotals[label] = (flt(net, 2), flt(vat, 2))
            ws.append([])

        # Reconciliation — the block the accountant's coloured footer was doing by
        # hand, stated as the arithmetic that produces the filed figures.
        fig = box_figures(plan, STANDARD_RATE)
        box = rule.get("target_box") or "1.2"
        amount, adjustment, vat_due = fig["amount"], fig["adjustment"], fig["vat"]

        ws.append([H("Reconciliation to box {0}".format(box),
                     "التسوية مع الخانة {0}".format(box))])
        ws[f"A{ws.max_row}"].font = Font(bold=True, size=12, color="1F4E5F")
        for (lab, val, bold) in (
            (H("Amount (declared this period)", "المبلغ (المصرح به هذه الفترة)"), amount, False),
            (H("Adjustment (credit notes)", "مبلغ التعديل (إشعارات دائنة)"), adjustment, False),
            (H("Taxable base", "الوعاء الضريبي"), fig["base"], False),
            (H("VAT at {0}%".format(int(STANDARD_RATE)), "الضريبة بنسبة {0}%".format(int(STANDARD_RATE))), vat_due, True),
        ):
            ws.append([lab, val])
            cc = ws.cell(row=ws.max_row, column=2)
            cc.number_format = "#,##0.00"; cc.alignment = Alignment(horizontal="right")
            cc.font = Font(bold=bold, color="11816F" if bold else "555555")
            ws.cell(row=ws.max_row, column=1).font = Font(bold=bold, italic=not bold)
            if bold:
                cc.fill = F_TOT; ws.cell(row=ws.max_row, column=1).fill = F_TOT

        # A cross-check, not a restatement: VAT summed off the invoices should equal
        # VAT recomputed from the base. A gap here means rounding on the invoices or
        # a non-standard rate inside a government supply, and the preparer needs to
        # see it rather than have one of the two figures quietly chosen for them.
        if abs(fig["variance"]) >= 0.01:
            ws.append([H("Check: VAT on the invoices differs from {0}% of the base by {1}"
                         .format(int(STANDARD_RATE), fig["variance"]),
                         "تنبيه: فرق بين ضريبة الفواتير و{0}% من الوعاء".format(int(STANDARD_RATE)))])
            ws[f"A{ws.max_row}"].font = Font(size=9, color="B3261E")

        # Ledger cross-check — the pull an accountant otherwise does by hand to
        # satisfy themselves the VAT account agrees with what the return says.
        # Reported, never reconciled away: a gap is information, and nothing here
        # can know which of the two numbers is the wrong one.
        checks = ledger_check(company, rule, from_date, to_date)
        if checks:
            ws.append([])
            ws.append([H("Ledger cross-check", "المطابقة مع دفتر الأستاذ")])
            ws[f"A{ws.max_row}"].font = Font(bold=True, size=12, color="1F4E5F")
            for chk in checks:
                ws.append(["{0} — {1}".format(_(chk["label"]), chk["account"]), chk["movement"]])
                cc = ws.cell(row=ws.max_row, column=2)
                cc.number_format = "#,##0.00"; cc.alignment = Alignment(horizontal="right")
                cc.font = Font(color="555555")
            deferred = subtotals.get(sections[2][0], (0.0, 0.0))[1]
            ws.append([H("Register: VAT carried forward", "السجل: ضريبة مرحّلة"), flt(deferred, 2)])
            cc = ws.cell(row=ws.max_row, column=2)
            cc.number_format = "#,##0.00"; cc.alignment = Alignment(horizontal="right")
            cc.font = Font(color="555555")
            ws.append([H("Compare the deferred account against the carried-forward figure above; "
                         "a difference is a posting the register cannot see.",
                         "قارن حساب الضريبة المؤجلة مع المبلغ المرحّل أعلاه؛ الفرق يعني قيداً لا يظهر في السجل.")])
            ws[f"A{ws.max_row}"].font = Font(size=9, italic=True, color="777777")

        if plan.get("ungrouped_customers"):
            ws.append([])
            ws.append([H("Customers with no customer group — GTPL scope undetermined: ",
                         "عملاء بدون مجموعة — نطاق النظام غير محدد: ")
                       + ", ".join(plan["ungrouped_customers"][:12])
                       + ("…" if len(plan["ungrouped_customers"]) > 12 else "")])
            ws[f"A{ws.max_row}"].font = Font(size=9, color="B3261E")
        autow(ws)

    for sheet_cfg in cfg.get("sheets") or []:
        stype = sheet_cfg.get("type")
        spec = _TYPES.get(stype)
        if not spec:
            continue
        opts = sheet_cfg.get("options") or {}
        if stype == "vat_return":
            data = vat_return(company, from_date, to_date)
            ws = new_sheet(sheet_cfg.get("title") or "VAT Return")
            head_block(ws, sheet_cfg.get("title") or _("VAT Return"))
            header_row(ws, ["#", H("Description", "الوصف"), H("Amount", "المبلغ"),
                            H("Adjustment", "التعديل"), H("VAT Amount", "مبلغ الضريبة")])
            for ln in (data.get("sales_lines") or []) + [data.get("box6") or {}] + \
                      (data.get("purchase_lines") or []) + [data.get("box12") or {}]:
                if not ln:
                    continue
                ws.append([ln.get("box"), ln.get("label"), flt(ln.get("amount"), 2),
                           flt(ln.get("adjustment"), 2), flt(ln.get("vat"), 2)])
                for j in (3, 4, 5):
                    ws.cell(row=ws.max_row, column=j).number_format = "#,##0.00"
            net = data.get("net") or {}
            ws.append([net.get("box"), _("Net VAT due"), None, None, flt(net.get("vat"), 2)])
            for c in ws[ws.max_row]:
                c.fill = F_TOT; c.font = Font(bold=True, color="11816F")
            ws.cell(row=ws.max_row, column=5).number_format = "#,##0.00"
            autow(ws)
        elif stype == "gst_sales_register":
            register_sheet(sheet_cfg, _gst_register_rows(company, from_date, to_date, "Sales Invoice"), spec)
        elif stype == "gst_purchase_register":
            register_sheet(sheet_cfg, _gst_register_rows(company, from_date, to_date, "Purchase Invoice"), spec)
        elif stype == "sales_register":
            register_sheet(sheet_cfg, _sales_rows(company, from_date, to_date, opts), spec)
        elif stype == "purchase_register":
            register_sheet(sheet_cfg, _purchase_rows(company, from_date, to_date, opts), spec)
        elif stype == "gtpl_register":
            gtpl_sheet(sheet_cfg, spec, opts)
        elif stype == "gl_ledger":
            rows, accounts = _gl_rows(company, from_date, to_date, opts)
            spec2 = dict(spec)
            ws_title = sheet_cfg.get("title") or spec["label"]
            register_sheet({**sheet_cfg, "title": ws_title}, rows, spec2)
            wb[wb.sheetnames[-1]].insert_rows(3)
            wb[wb.sheetnames[-1]]["A3"] = _("Accounts: ") + ", ".join(accounts[:6]) + ("…" if len(accounts) > 6 else "")
            wb[wb.sheetnames[-1]]["A3"].font = Font(size=9, color="777777")

    if not wb.sheetnames:
        frappe.throw(_("The pack produced no sheets — check its configuration."))

    buf = io.BytesIO()
    wb.save(buf)
    frappe.local.response.filename = f"{frappe.scrub(pack['title'])}-{from_date}-to-{to_date}.xlsx"
    frappe.local.response.filecontent = buf.getvalue()
    frappe.local.response.type = "binary"


@frappe.whitelist()
def list_gl_accounts(company=None):
    """Leaf accounts for the pack designer's ledger picker."""
    company = company or _default_company()
    if not frappe.has_permission("Account", "read"):
        return []
    return frappe.get_all("Account", filters={"company": company, "is_group": 0},
                          fields=["name", "account_name as label"],
                          order_by="name", limit_page_length=0)
