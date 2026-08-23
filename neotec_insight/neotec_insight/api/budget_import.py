# Copyright (c) 2026, Neotec Integrated Solution
# Budget Import (v2.37.0) — bulk budgeting for dimension books. One template,
# all cost centers (or departments/projects/custom values): download it,
# fill amounts, upload back. The engine validates before writing anything and
# creates/updates one budget book per dimension value plus its cells.
from __future__ import annotations

import io
import json

import frappe
from frappe import _
from frappe.utils import flt

_DIM_DOCTYPE = {"cost_center": "Cost Center", "department": "Department", "project": "Project"}


def _source_rows(report):
    doc = frappe.get_doc("Insight Report Definition", report)
    definition = json.loads(doc.definition_json or "{}")
    return doc, [r for r in (definition.get("rows") or [])
                 if r.get("kind") == "source" and not r.get("hidden")]


@frappe.whitelist()
def budget_import_template(report=None, fiscal_year=None, dimension_type="cost_center", values=None):
    """Styled xlsx template: Dimension Value | Row Key | Row Label | M1..M12.
    One block per dimension value, rows = the report's source rows."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    doc, rows = _source_rows(report)
    if not frappe.has_permission("Insight Budget Book", "create"):
        frappe.throw(_("Not permitted."))
    if isinstance(values, str):
        values = [v.strip() for v in values.split("|") if v.strip()]
    if not values:
        dt = _DIM_DOCTYPE.get(dimension_type)
        if dt:
            flt_ = {"company": doc.company} if dt != "Project" else {}
            if dt == "Cost Center":
                flt_["is_group"] = 0
            values = frappe.get_all(dt, filters=flt_, pluck="name", limit_page_length=200)
    if not values:
        frappe.throw(_("No dimension values found — pass them explicitly."))

    wb = Workbook(); ws = wb.active; ws.title = "Budget"
    ws.append([f"Report: {doc.report_name}", f"FY: {fiscal_year}", f"Dimension: {dimension_type}"])
    ws["A1"].font = Font(bold=True)
    head = ["Dimension Value", "Row Key", "Row Label"] + [f"M{i}" for i in range(1, 13)]
    ws.append(head)
    for c in ws[2]:
        c.fill = PatternFill("solid", fgColor="2A2440"); c.font = Font(bold=True, color="FFFFFF")
    for v in values:
        for r in rows:
            ws.append([v, r.get("key"), r.get("label")] + [0] * 12)
    ws.column_dimensions["A"].width = 34; ws.column_dimensions["C"].width = 40
    buf = io.BytesIO(); wb.save(buf)
    frappe.local.response.filename = f"budget-template-{frappe.scrub(doc.report_name)}-fy{fiscal_year}.xlsx"
    frappe.local.response.filecontent = buf.getvalue()
    frappe.local.response.type = "binary"


@frappe.whitelist()
def budget_import(report=None, fiscal_year=None, dimension_type="cost_center",
                  custom_dimension_fieldname=None, rows=None):
    """rows = [{value, row_key, months:{'0'..'11': amount}}] parsed client-side.
    Validates everything first; then one book per value (created if missing,
    DRAFT) and upserted cells. M1..M12 in the template = months 0..11 of the
    company's fiscal year (Apr-first for India, Jan-first for KSA)."""
    if not frappe.has_permission("Insight Budget Book", "create"):
        frappe.throw(_("Not permitted."))
    doc, src_rows = _source_rows(report)
    valid_keys = {r.get("key") for r in src_rows}
    fiscal_year = int(fiscal_year)
    if isinstance(rows, str):
        rows = json.loads(rows or "[]")

    warnings, clean = [], []
    dt = _DIM_DOCTYPE.get(dimension_type)
    seen_values = set()
    for r in rows or []:
        value = (r.get("value") or "").strip()
        key = (r.get("row_key") or "").strip()
        if not value:
            continue
        if key not in valid_keys:
            warnings.append(_("Unknown row key skipped: {0}").format(key))
            continue
        if dt and value not in seen_values and not frappe.db.exists(dt, value):
            warnings.append(_("{0} not found in ERP: {1} (book still created)").format(_(dt), value))
        seen_values.add(value)
        months = {}
        for m, amt in (r.get("months") or {}).items():
            try:
                mi = int(m)
                if 0 <= mi <= 11:
                    months[mi] = flt(amt)
            except Exception:
                continue
        clean.append({"value": value, "key": key, "months": months})

    if not clean:
        frappe.throw(_("Nothing importable — check row keys and values."))

    books_created, cells = 0, 0
    book_cache = {}
    for value in {c["value"] for c in clean}:
        existing = frappe.get_all("Insight Budget Book",
                                  filters={"report": doc.name, "fiscal_year": fiscal_year,
                                           "dimension_type": dimension_type,
                                           "dimension_value": value}, pluck="name")
        if existing:
            book_cache[value] = existing[0]
        else:
            b = frappe.new_doc("Insight Budget Book")
            b.report = doc.name
            b.fiscal_year = fiscal_year
            b.dimension_type = dimension_type
            b.dimension_value = value
            if dimension_type == "custom" and custom_dimension_fieldname:
                b.custom_dimension_fieldname = custom_dimension_fieldname
            b.status = "draft"
            # v2.37.2 — slug (docname) and label are derived, which is why the
            # template has no book-name column: FY + dimension value ARE the
            # identity of a book.
            dim_label = {"cost_center": "Cost Center", "department": "Department",
                         "project": "Project", "custom": custom_dimension_fieldname or "Dimension"}
            b.label = f"FY{fiscal_year} \u00b7 {dim_label.get(dimension_type, dimension_type)}: {value}"
            base = frappe.scrub(f"{doc.slug or doc.name}-fy{fiscal_year}-{dimension_type}-{value}")[:120]
            slug = base
            i = 2
            while frappe.db.exists("Insight Budget Book", slug):
                slug = f"{base}-{i}"
                i += 1
            b.slug = slug
            b.insert(ignore_permissions=True)
            book_cache[value] = b.name
            books_created += 1

    for c in clean:
        book = book_cache[c["value"]]
        for m, amt in c["months"].items():
            name = frappe.get_all("Insight Budget Cell",
                                  filters={"book": book, "row_key": c["key"], "month": m},
                                  pluck="name")
            if name:
                frappe.db.set_value("Insight Budget Cell", name[0], "amount", amt)
            else:
                frappe.get_doc({"doctype": "Insight Budget Cell", "book": book,
                                "row_key": c["key"], "month": m,
                                "amount": amt}).insert(ignore_permissions=True)
            cells += 1
    frappe.db.commit()
    return {"books_created": books_created, "books_touched": len(book_cache),
            "cells_written": cells, "warnings": warnings[:30]}
