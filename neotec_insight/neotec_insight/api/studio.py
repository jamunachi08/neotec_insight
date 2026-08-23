# Copyright (c) 2026, Neotec Integrated Solution
# Report Studio — a no-code, AI-assisted reporting/BI layer over any DocType.
# Fully isolated from the financial engine (report.py / execution.py).
from __future__ import annotations

import json

import frappe
from frappe import _

# Field types that make sense as report columns.
_REPORTABLE = {
    "Data", "Select", "Link", "Dynamic Link", "Small Text", "Text", "Long Text",
    "Read Only", "Int", "Float", "Currency", "Percent", "Check", "Date", "Datetime",
    "Time", "Duration", "Rating", "Code",
}
_NUMERIC = {"Int", "Float", "Currency", "Percent"}

# Standard fields always available on a DocType.
_STD_FIELDS = [
    {"fieldname": "name", "label": "ID", "fieldtype": "Data"},
    {"fieldname": "owner", "label": "Created By", "fieldtype": "Link"},
    {"fieldname": "creation", "label": "Created On", "fieldtype": "Datetime"},
    {"fieldname": "modified", "label": "Last Updated", "fieldtype": "Datetime"},
]


def _require_read() -> None:
    """Refuse a financial read from a user with no ledger access.

    `@frappe.whitelist()` requires a login, not a role, so without this these
    endpoints were callable over `/api/method/...` by any authenticated user,
    including portal users with no business seeing the ledger. Reading GL Entry
    is the right test: ERPNext already restricts it to the accounts roles, so
    this inherits the site's own configuration rather than inventing a second
    permission model.
    """
    if not frappe.has_permission("GL Entry", "read"):
        frappe.throw(
            _("You are not permitted to view financial data."),
            frappe.PermissionError,
        )


def _can_read(doctype: str) -> bool:
    try:
        if frappe.has_permission(doctype, "read"):
            return True
        # Child tables carry no permissions of their own — defer to the parent.
        if frappe.get_meta(doctype).istable:
            parent = _child_parent(doctype)
            return bool(parent and frappe.has_permission(parent, "read"))
        return False
    except Exception:
        return False


def _child_parent(child_doctype: str):
    """Primary parent DocType of a child table (a child like 'Sales Taxes and
    Charges' can belong to several parents; we pick the highest-priority one and
    scope queries to it via parenttype)."""
    parents = frappe.get_all(
        "DocField", filters={"fieldtype": "Table", "options": child_doctype},
        fields=["parent"], limit_page_length=20)
    names = sorted({p["parent"] for p in parents})
    if not names:
        return None
    priority = ["Sales Invoice", "Purchase Invoice", "Sales Order", "Purchase Order",
                "Delivery Note", "Purchase Receipt", "Journal Entry", "Payment Entry",
                "Quotation", "Stock Entry"]
    names.sort(key=lambda n: (priority.index(n) if n in priority else 999, n))
    return names[0]



def _require_write(doctype: str) -> None:
    """Refuse a write from a user who lacks permission on the doctype.

    Frappe checks permissions inside `doc.save()`, but these endpoints use
    `ignore_permissions=True`, `frappe.db.set_value` or raw SQL — all of which
    bypass that check. Whitelisted and unguarded, they were callable directly
    over `/api/method/...` by any authenticated user, including a read-only
    one. The guard restores the check the bypass removed.
    """
    if not frappe.has_permission(doctype, "write"):
        frappe.throw(
            _("You are not permitted to modify {0}.").format(_(doctype)),
            frappe.PermissionError,
        )


@frappe.whitelist()
def list_sources(search=""):
    """Documents (DocTypes) the user may report on — parents AND child tables.

    Child tables (e.g. Sales Invoice Item) are first-class sources: they carry
    the line-level detail (items, quantities, rates) that parent-level reporting
    can't reach. Each is labelled with its parent and read-permission is checked
    against the parent DocType."""
    _require_read()
    search = (search or "").strip()
    or_filters = None
    if search:
        or_filters = [["name", "like", f"%{search}%"], ["module", "like", f"%{search}%"]]
    rows = frappe.get_all(
        "DocType", filters={"issingle": 0}, or_filters=or_filters,
        fields=["name", "module", "istable"],
        order_by="modified desc", limit_page_length=600,
    )
    out = []
    for r in rows:
        if r.get("istable"):
            parent = _child_parent(r["name"])
            if not parent or not _can_read(parent):
                continue
            out.append({"name": r["name"], "label": _(r["name"]),
                        "module": r.get("module") or "", "is_child": 1,
                        "parent_doctype": parent})
        else:
            if not _can_read(r["name"]):
                continue
            out.append({"name": r["name"], "label": _(r["name"]), "module": r.get("module") or ""})
        if len(out) >= 120:
            break
    # Sort common business docs (and their item tables) to the top.
    priority = ["Sales Invoice", "Sales Invoice Item", "Purchase Invoice", "Purchase Invoice Item",
                "Payment Entry", "Journal Entry", "Sales Order", "Sales Order Item",
                "Purchase Order", "Delivery Note", "Customer", "Supplier",
                "Item", "Employee", "GL Entry", "Stock Ledger Entry"]
    out.sort(key=lambda d: (priority.index(d["name"]) if d["name"] in priority else 999, d["name"]))
    return out


@frappe.whitelist()
def field_values(doctype=None, field=None, search=""):
    """Distinct value suggestions for a filter field — Link targets or the
    actual values present, so users can pick instead of typing."""
    _require_read()
    if not doctype or not field or not _can_read(doctype):
        return []
    meta = frappe.get_meta(doctype)
    f = meta.get_field(field)
    search = (search or "").strip()
    if f and f.fieldtype == "Select" and f.options:
        opts = [o for o in (f.options or "").split("\n") if o]
        return [{"value": o, "label": o} for o in opts if not search or search.lower() in o.lower()][:50]
    if f and f.fieldtype == "Link" and f.options and _can_read(f.options):
        flt = [[f.options, "name", "like", f"%{search}%"]] if search else None
        rows = frappe.get_list(f.options, filters=flt, fields=["name"], limit_page_length=25, order_by="modified desc")
        return [{"value": r["name"], "label": r["name"]} for r in rows]
    # Fallback: distinct values present on the base doctype.
    try:
        vals = frappe.get_all(doctype, fields=[f"`{field}` as v"], group_by=field,
                              limit_page_length=50, order_by=field)
        return [{"value": v["v"], "label": str(v["v"])} for v in vals if v.get("v") not in (None, "")]
    except Exception:
        return []


@frappe.whitelist()
def list_fields(doctype=None):
    """Reportable fields for a DocType, plus standard fields."""
    _require_read()
    if not doctype or not _can_read(doctype):
        return []
    meta = frappe.get_meta(doctype)
    fields = []
    for f in meta.fields:
        if f.fieldtype in _REPORTABLE and f.fieldname:
            fields.append({
                "fieldname": f.fieldname,
                "label": _(f.label) if f.label else f.fieldname,
                "fieldtype": f.fieldtype,
                "options": f.options or "",
                "numeric": f.fieldtype in _NUMERIC,
            })
    # Append standard fields not already present.
    have = {f["fieldname"] for f in fields}
    std = list(_STD_FIELDS)
    if meta.istable:
        # Child rows: expose the parent document reference and row position.
        std = [{"fieldname": "parent", "label": "Parent Document", "fieldtype": "Data"},
               {"fieldname": "idx", "label": "Row #", "fieldtype": "Int"}] + std
    for sf in std:
        if sf["fieldname"] not in have:
            fields.append({**sf, "options": "", "numeric": sf["fieldtype"] == "Int"})
    return fields


def _parse_user_date(val):
    """Accept the date formats humans actually type — 01012026, 010126,
    01/01/2026, 01-01-26, 01.01.2026, or ISO — and return ISO yyyy-mm-dd.
    Day-first (KSA/EU convention) unless the first token is 4 digits (ISO).
    Anything unparseable is returned untouched so Frappe raises its own error."""
    import re as _re
    from datetime import date as _date
    sv = str(val or "").strip()
    if not sv:
        return val
    try:
        if _re.fullmatch(r"\d{8}", sv):          # ddmmyyyy
            d, m, y = int(sv[0:2]), int(sv[2:4]), int(sv[4:8])
        elif _re.fullmatch(r"\d{6}", sv):        # ddmmyy
            d, m, y = int(sv[0:2]), int(sv[2:4]), 2000 + int(sv[4:6])
        else:
            parts = _re.split(r"[/.\-\s]+", sv)
            if len(parts) != 3 or not all(p.isdigit() for p in parts):
                return val
            if len(parts[0]) == 4:                # yyyy-mm-dd (ISO)
                y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
            else:                                 # dd-mm-yyyy / dd-mm-yy
                d, m = int(parts[0]), int(parts[1])
                y = int(parts[2]) + (2000 if len(parts[2]) == 2 else 0)
        return _date(y, m, d).isoformat()
    except Exception:
        return val


def _coerce_filters(doctype, raw_filters):
    """Translate UI filters into safe Frappe filter tuples. Date/Datetime
    fields get human-format normalisation (ddmmyyyy, dd/mm/yyyy, …)."""
    out = []
    meta = frappe.get_meta(doctype)
    valid = {f.fieldname for f in meta.fields} | {"name", "owner", "creation", "modified",
                                                   "parent", "parenttype", "parentfield", "idx", "docstatus"}
    date_fields = {f.fieldname for f in meta.fields if f.fieldtype in ("Date", "Datetime")} | {"creation", "modified"}
    for fl in (raw_filters or []):
        field = fl.get("field")
        op = (fl.get("op") or "=").strip()
        val = fl.get("value")
        if field not in valid:
            continue
        if field in date_fields:
            val = [_parse_user_date(v) for v in val] if isinstance(val, list) else _parse_user_date(val)
        if op in ("is set", "is not set"):
            out.append([doctype, field, "is", "set" if op == "is set" else "not set"])
        elif op == "in":
            vals = val if isinstance(val, list) else [v.strip() for v in str(val).split(",") if v.strip()]
            out.append([doctype, field, "in", vals])
        elif op == "between":
            vals = val if isinstance(val, list) else [v.strip() for v in str(val).split(",")]
            out.append([doctype, field, "between", vals[:2]])
        elif op == "like":
            out.append([doctype, field, "like", f"%{val}%"])
        else:
            out.append([doctype, field, op, val])
    return out


@frappe.whitelist()
def list_link_fields(doctype=None):
    """Link fields on a DocType — the bridges to combine other documents."""
    _require_read()
    if not doctype or not _can_read(doctype):
        return []
    meta = frappe.get_meta(doctype)
    out = []
    # Child table → parent join: the child's `parent` column holds the parent
    # document name, so it behaves exactly like a Link for our join machinery.
    if meta.istable:
        parent = _child_parent(doctype)
        if parent and _can_read(parent):
            out.append({"link_field": "parent",
                        "label": _("Parent ({0})").format(_(parent)),
                        "target_doctype": parent})
    for f in meta.fields:
        if f.fieldtype in ("Link", "Dynamic Link") and f.options and f.fieldtype == "Link":
            if _can_read(f.options):
                out.append({"link_field": f.fieldname,
                            "label": _(f.label) if f.label else f.fieldname,
                            "target_doctype": f.options})
    return out


import ast as _ast

_ALLOWED_FUNCS = {"abs": abs, "round": round, "min": min, "max": max}


def _safe_eval(expr, scope):
    """Evaluate a numeric formula safely (no eval/exec). Supports + - * / % **,
    parentheses, unary minus, numbers, field references, and abs/round/min/max."""
    try:
        node = _ast.parse(str(expr), mode="eval").body
    except Exception:
        return None

    def ev(n):
        if isinstance(n, _ast.Constant):
            return n.value if isinstance(n.value, (int, float)) else 0
        if isinstance(n, _ast.Name):
            v = scope.get(n.id, 0)
            try:
                return float(v) if v not in (None, "") else 0
            except Exception:
                return 0
        if isinstance(n, _ast.BinOp):
            a, b = ev(n.left), ev(n.right)
            op = n.op
            if isinstance(op, _ast.Add):
                return a + b
            if isinstance(op, _ast.Sub):
                return a - b
            if isinstance(op, _ast.Mult):
                return a * b
            if isinstance(op, _ast.Div):
                return a / b if b else 0
            if isinstance(op, _ast.Mod):
                return a % b if b else 0
            if isinstance(op, _ast.Pow):
                return a ** b
            raise ValueError("op")
        if isinstance(n, _ast.UnaryOp) and isinstance(n.op, (_ast.USub, _ast.UAdd)):
            v = ev(n.operand)
            return -v if isinstance(n.op, _ast.USub) else v
        if isinstance(n, _ast.Call) and isinstance(n.func, _ast.Name) and n.func.id in _ALLOWED_FUNCS:
            return _ALLOWED_FUNCS[n.func.id](*[ev(a) for a in n.args])
        raise ValueError("unsupported")

    try:
        return ev(node)
    except Exception:
        return None


def _run_pivot(doctype, ftype, flabel, config, pivot):
    """Build a cross-tab: Row field × Column field, aggregating a measure."""
    rowf, colf, valf = pivot.get("row"), pivot.get("column"), pivot.get("value")
    agg = (pivot.get("agg") or "sum").lower()
    if rowf not in ftype:
        frappe.throw(_("Pivot needs a valid Row field."))
    single = not (colf and colf in ftype)
    if agg != "count":
        if not valf or valf not in ftype:
            frappe.throw(_("Pick a Value field for the pivot (or use the Count aggregation)."))
    else:
        valf = valf if (valf and valf in ftype) else None
    limit = min(int(config.get("limit") or 2000), 20000)
    fetch = [rowf] + ([colf] if not single else []) + ([valf] if valf else [])
    fetch_filters = _coerce_filters(doctype, config.get("filters"))
    kwargs = {}
    if frappe.get_meta(doctype).istable:
        cp = _child_parent(doctype)
        if cp:
            fetch_filters.append([doctype, "parenttype", "=", cp])
            kwargs["parent_doctype"] = cp
    rows = frappe.get_list(doctype, filters=fetch_filters,
                           fields=list(dict.fromkeys(fetch)), limit_page_length=limit, **kwargs)

    from collections import defaultdict
    cells = defaultdict(lambda: defaultdict(list))
    row_vals, col_vals, all_vals = defaultdict(list), defaultdict(list), []
    rkeys, ckeys, rseen, cseen = [], [], set(), set()
    for r in rows:
        rk = r.get(rowf) or "—"
        v = r.get(valf) if valf else 1
        if rk not in rseen: rseen.add(rk); rkeys.append(rk)
        row_vals[rk].append(v); all_vals.append(v)
        if not single:
            ck = r.get(colf) or "—"
            if ck not in cseen: cseen.add(ck); ckeys.append(ck)
            cells[rk][ck].append(v); col_vals[ck].append(v)
    rkeys.sort(key=lambda x: str(x)); ckeys.sort(key=lambda x: str(x))

    def _agg(vals):
        if agg == "count":
            return len(vals)
        nums = [float(v) for v in vals if v not in (None, "")]
        if not nums:
            return 0
        if agg == "avg":
            return sum(nums) / len(nums)
        if agg == "min":
            return min(nums)
        if agg == "max":
            return max(nums)
        return sum(nums)

    return {
        "pivot": {
            "row_field": rowf, "row_label": flabel.get(rowf, rowf),
            "col_field": colf, "col_label": flabel.get(colf, colf) if not single else "",
            "value_label": (flabel.get(valf, valf) if valf else _("Count")), "agg": agg,
            "single": single,
            "columns": [] if single else [str(c) for c in ckeys],
            "rows": [{"key": str(rk),
                      "cells": {} if single else {str(ck): _agg(cells[rk][ck]) for ck in ckeys},
                      "total": _agg(row_vals[rk])} for rk in rkeys],
            "col_totals": {} if single else {str(ck): _agg(col_vals[ck]) for ck in ckeys},
            "grand_total": _agg(all_vals),
        },
        "row_count": len(rows),
    }


@frappe.whitelist()
def run_query(config=None):
    """Execute a Studio report config and return rows (+ optional grouping).

    config = {
      doctype,
      columns:[fieldname | "link_field.target_field" ...],   # dotted = joined doc
      filters:[{field,op,value}], group_by, sort:[{field,dir}], limit,
      calculated:[{key,label,formula,decimals}],             # formula columns
      col_modes:{ colKey: "running"|"pct_total" }            # per-column "show as"
    }
    Uses frappe.get_list (permission-safe) for the base doc and batched
    frappe.get_all lookups for each linked document.
    """
    _require_read()
    if isinstance(config, str):
        config = json.loads(config or "{}")
    doctype = config.get("doctype")
    if not doctype:
        frappe.throw(_("Pick a document first."))
    if not _can_read(doctype):
        frappe.throw(_("You don't have permission to read {0}.").format(doctype))

    meta = frappe.get_meta(doctype)
    ftype = {f.fieldname: f.fieldtype for f in meta.fields}
    flabel = {f.fieldname: (_(f.label) if f.label else f.fieldname) for f in meta.fields}
    link_opts = {f.fieldname: f.options for f in meta.fields if f.fieldtype == "Link"}
    for sf in _STD_FIELDS:
        ftype.setdefault(sf["fieldname"], sf["fieldtype"])
        flabel.setdefault(sf["fieldname"], sf["label"])

    # Child table as source: `parent` behaves as a Link to the parent DocType so
    # dotted columns like parent.customer join parent fields; queries are scoped
    # to this parent via parenttype (shared child tables serve several parents)
    # and run with parent_doctype so Frappe applies the PARENT's permissions.
    child_parent = _child_parent(doctype) if meta.istable else None
    if child_parent:
        link_opts["parent"] = child_parent
        for fn, lb, ft in (("parent", f"Parent ({child_parent})", "Data"),
                           ("parenttype", "Parent Type", "Data"),
                           ("idx", "Row #", "Int"), ("docstatus", "Doc Status", "Int")):
            ftype.setdefault(fn, ft)
            flabel.setdefault(fn, lb)

    pivot = config.get("pivot") or None
    if pivot and pivot.get("row") and pivot.get("column") and pivot.get("value"):
        return _run_pivot(doctype, ftype, flabel, config, pivot)

    raw_cols = config.get("columns") or []
    plain_cols, dotted_cols = [], []
    for c in raw_cols:
        if "." in c:
            lf, tf = c.split(".", 1)
            if lf in link_opts and _can_read(link_opts[lf]):
                dotted_cols.append((c, lf, tf, link_opts[lf]))
        elif c in ftype:
            plain_cols.append(c)
    columns = plain_cols + [d[0] for d in dotted_cols]
    if not columns:
        columns = ["name"]; plain_cols = ["name"]

    group_by = config.get("group_by") or None
    if group_by and group_by not in ftype:
        group_by = None

    fetch = list(dict.fromkeys(plain_cols + ([group_by] if group_by else [])
                               + [d[1] for d in dotted_cols]))
    sort = config.get("sort") or []
    order_by = ", ".join(f"`{s['field']}` {('desc' if s.get('dir') == 'desc' else 'asc')}"
                         for s in sort if s.get("field") in ftype) or None
    if group_by and not order_by:
        order_by = f"`{group_by}` asc"

    limit = min(int(config.get("limit") or 1000), 5000)
    fetch_filters = _coerce_filters(doctype, config.get("filters"))
    kwargs = {}
    if child_parent:
        # Scope shared child tables to the primary parent and inherit its perms.
        fetch_filters.append([doctype, "parenttype", "=", child_parent])
        kwargs["parent_doctype"] = child_parent
    rows = frappe.get_list(
        doctype, filters=fetch_filters,
        fields=fetch, order_by=order_by, limit_page_length=limit, **kwargs,
    )

    # ── Joins: batch-fetch each linked document and attach its fields. ──
    col_meta = {}  # colKey -> {label,type,numeric}
    for c in plain_cols:
        col_meta[c] = {"label": flabel.get(c, c), "type": ftype.get(c, "Data"),
                       "numeric": ftype.get(c) in _NUMERIC}
    by_link = {}
    for (ckey, lf, tf, target) in dotted_cols:
        by_link.setdefault((lf, target), []).append((ckey, tf))
    for (lf, target), pairs in by_link.items():
        tmeta = frappe.get_meta(target)
        ttype = {f.fieldname: f.fieldtype for f in tmeta.fields}
        tlabel = {f.fieldname: (_(f.label) if f.label else f.fieldname) for f in tmeta.fields}
        tlabel.setdefault("name", "ID")
        tfields = list(dict.fromkeys(["name"] + [tf for _ck, tf in pairs if tf in ttype or tf == "name"]))
        keys = list({r.get(lf) for r in rows if r.get(lf)})
        lookup = {}
        if keys:
            for trow in frappe.get_all(target, filters={"name": ["in", keys]}, fields=tfields,
                                       limit_page_length=0):
                lookup[trow["name"]] = trow
        for r in rows:
            src = lookup.get(r.get(lf), {})
            for (ckey, tf) in pairs:
                r[ckey] = src.get(tf)
        for (ckey, tf) in pairs:
            col_meta[ckey] = {"label": f"{flabel.get(lf, lf)} · {tlabel.get(tf, tf)}",
                              "type": ttype.get(tf, "Data"),
                              "numeric": ttype.get(tf) in _NUMERIC}

    # ── Calculated (formula) columns, computed per row over numeric scope. ──
    calculated = config.get("calculated") or []
    calc_keys = []
    for calc in calculated:
        key = (calc.get("key") or "").strip()
        if not key or not key.isidentifier():
            continue
        formula = calc.get("formula") or "0"
        dec = int(calc.get("decimals", 2))
        for r in rows:
            scope = {k: r.get(k) for k in (plain_cols + [d[0] for d in dotted_cols] + calc_keys)}
            val = _safe_eval(formula, scope)
            r[key] = round(val, dec) if isinstance(val, (int, float)) else None
        col_meta[key] = {"label": calc.get("label") or key, "type": "Float", "numeric": True}
        columns.append(key)
        calc_keys.append(key)

    num_cols = [c for c in columns if col_meta.get(c, {}).get("numeric")]

    def _sum(rs):
        return {c: sum(float(r.get(c) or 0) for r in rs) for c in num_cols}

    grand = _sum(rows)

    # ── Per-column "show as": running total / % of total (display transform). ──
    col_modes = config.get("col_modes") or {}
    for ckey, mode in col_modes.items():
        if ckey not in num_cols:
            continue
        if mode == "pct_total":
            tot = grand.get(ckey) or 0
            for r in rows:
                r[ckey] = round((float(r.get(ckey) or 0) / tot * 100), 2) if tot else 0
        elif mode == "running":
            run = 0.0
            for r in rows:
                run += float(r.get(ckey) or 0)
                r[ckey] = round(run, 2)

    out = {
        "columns": [{"field": c, **col_meta.get(c, {"label": c, "type": "Data", "numeric": False}),
                     "mode": col_modes.get(c)} for c in columns],
        "group_by": group_by,
        "group_label": flabel.get(group_by) if group_by else None,
        "row_count": len(rows),
        "grand_total": grand,
    }
    if group_by:
        # Optional Sales/Returns split: rows flagged as returns are listed and
        # subtotalled separately; the group subtotal = Sales − Returns (net).
        ret = config.get("returns") or None
        rfield = (ret or {}).get("field") if ret else None
        rvalue = str((ret or {}).get("value", "1")) if ret else None
        split_returns = bool(rfield and rfield in ftype)

        def _is_return(r):
            return split_returns and str(r.get(rfield)) == rvalue

        groups = {}
        for r in rows:
            groups.setdefault(r.get(group_by) or "—", []).append(r)
        out["split_returns"] = split_returns
        glist = []
        for k, rs in groups.items():
            g = {"key": k, "count": len(rs), "subtotal": _sum(rs)}
            if split_returns:
                sales = [r for r in rs if not _is_return(r)]
                rets = [r for r in rs if _is_return(r)]
                ssub, rsub = _sum(sales), _sum(rets)
                g.update({
                    "sales_rows": sales, "return_rows": rets,
                    "sales_subtotal": ssub, "return_subtotal": rsub,
                    "net_subtotal": {c: (ssub.get(c, 0) + rsub.get(c, 0)) for c in num_cols},
                })
            else:
                g["rows"] = rs
            glist.append(g)
        out["groups"] = glist
    else:
        out["rows"] = rows
    return out


@frappe.whitelist()
def ai_build(doctype=None, prompt=None):
    """AI magic: turn a plain-language request into a Studio query config.

    Sends the DocType's field list + the user's sentence to the configured LLM
    (Insight AI Settings) and asks for STRICT JSON: {columns, filters, group_by,
    sort}. Returns the parsed config the builder can apply directly.
    """
    _require_read()
    if not doctype or not prompt:
        frappe.throw(_("Pick a document and describe what you want."))
    fields = list_fields(doctype)
    field_catalog = "\n".join(f"- {f['fieldname']} ({f['fieldtype']}): {f['label']}" for f in fields)

    s = frappe.get_doc("Insight AI Settings")
    if not getattr(s, "ai_enabled", 0) or not getattr(s, "ai_endpoint", None):
        frappe.throw(_("AI is not configured. Enable it in Insight AI Settings."))

    system = (
        "You are a report-building assistant. Given a DocType's fields and a user's "
        "request, output ONLY a strict JSON object (no prose, no markdown) with keys: "
        "columns (array of fieldnames to show), filters (array of {field, op, value} "
        "where op is one of =,!=,>,<,>=,<=,like,in,between,is set,is not set), "
        "group_by (a fieldname or null), sort (array of {field, dir} dir=asc|desc). "
        "Only use fieldnames from the provided list. Keep it minimal and correct."
    )
    user = f"DocType: {doctype}\nAvailable fields:\n{field_catalog}\n\nRequest: {prompt}\n\nReturn JSON only."

    headers = {"Content-Type": "application/json"}
    api_key = s.get_password("ai_api_key") if s.ai_api_key else None
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"
    endpoint = s.ai_endpoint.rstrip("/") + "/chat/completions"
    payload = {"model": s.ai_model or "llama3",
               "messages": [{"role": "system", "content": system},
                            {"role": "user", "content": user}],
               "stream": False, "temperature": 0.1}
    import requests
    try:
        resp = requests.post(endpoint, headers=headers, json=payload, timeout=120)
        resp.raise_for_status()
        data = resp.json()
        text = data.get("choices", [{}])[0].get("message", {}).get("content", "")
    except Exception as e:  # noqa: BLE001
        frappe.log_error(frappe.get_traceback(), "Studio · ai_build")
        frappe.throw(_("AI request failed: {0}").format(str(e)[:200]))

    # Extract the JSON object from the response.
    text = (text or "").strip()
    start, end = text.find("{"), text.rfind("}")
    if start >= 0 and end > start:
        text = text[start:end + 1]
    try:
        cfg = json.loads(text)
    except Exception:
        frappe.throw(_("The AI returned something I couldn't parse. Try rephrasing."))
    # Validate fieldnames.
    valid = {f["fieldname"] for f in fields}
    cfg["columns"] = [c for c in (cfg.get("columns") or []) if c in valid]
    cfg["filters"] = [f for f in (cfg.get("filters") or []) if f.get("field") in valid]
    if cfg.get("group_by") not in valid:
        cfg["group_by"] = None
    cfg["sort"] = [s for s in (cfg.get("sort") or []) if s.get("field") in valid]
    cfg["doctype"] = doctype
    return cfg


@frappe.whitelist()
def save_report(report=None):
    """Create/update a saved Studio Report."""
    _require_write("Insight Report Definition")
    if isinstance(report, str):
        report = json.loads(report or "{}")
    title = (report.get("title") or "").strip()
    if not title:
        frappe.throw(_("Give the report a title."))
    slug = frappe.scrub(title)
    if frappe.db.exists("Studio Report", slug):
        doc = frappe.get_doc("Studio Report", slug)
    else:
        doc = frappe.new_doc("Studio Report")
        doc.slug = slug
    doc.title = title
    doc.description = report.get("description") or ""
    doc.config_json = json.dumps(report.get("config") or {})
    doc.save(ignore_permissions=True)
    frappe.db.commit()
    return {"slug": doc.slug, "title": doc.title}


@frappe.whitelist()
def list_reports():
    _require_read()
    return frappe.get_all("Studio Report", fields=["slug", "title", "description", "modified"],
                          order_by="modified desc", limit_page_length=200)


@frappe.whitelist()
def load_report(slug=None):
    _require_read()
    if not slug or not frappe.db.exists("Studio Report", slug):
        frappe.throw(_("Report not found."))
    doc = frappe.get_doc("Studio Report", slug)
    return {"slug": doc.slug, "title": doc.title, "description": doc.description,
            "config": json.loads(doc.config_json or "{}")}


# ────────────────────────────────────────────────────────────────────────────
# Time Intelligence — one-click MTD/QTD/YTD/YoY/rolling-12M for any measure.
# ────────────────────────────────────────────────────────────────────────────

def _fiscal_year_start(as_of):
    """Fiscal-year start for a date — ERPNext Fiscal Year when available,
    calendar 1 January otherwise (Insight already supports KSA Jan-start)."""
    from frappe.utils import getdate
    d = getdate(as_of)
    try:
        fy = frappe.get_all("Fiscal Year",
                            filters={"year_start_date": ["<=", d], "year_end_date": [">=", d],
                                     "disabled": 0},
                            fields=["year_start_date"], limit=1)
        if fy:
            return getdate(fy[0]["year_start_date"])
    except Exception:
        pass
    return d.replace(month=1, day=1)


def _ti_windows(as_of):
    """The comparison windows, each as (key, label, start, end)."""
    from dateutil.relativedelta import relativedelta
    from frappe.utils import getdate
    d = getdate(as_of)
    m_start = d.replace(day=1)
    pm_end = m_start - relativedelta(days=1)
    pm_start = pm_end.replace(day=1)
    pm_cut = min(d.day, pm_end.day)
    q_month = ((d.month - 1) // 3) * 3 + 1
    q_start = d.replace(month=q_month, day=1)
    fy_start = _fiscal_year_start(d)
    py_d = d - relativedelta(years=1)
    py_fy_start = fy_start - relativedelta(years=1)
    r12_start = d - relativedelta(months=12) + relativedelta(days=1)
    pr12_end = r12_start - relativedelta(days=1)
    pr12_start = pr12_end - relativedelta(months=12) + relativedelta(days=1)
    return [
        ("mtd",    "MTD",            m_start,     d),
        ("pmtd",   "Prior MTD",      pm_start,    pm_start.replace(day=pm_cut)),
        ("qtd",    "QTD",            q_start,     d),
        ("ytd",    "YTD",            fy_start,    d),
        ("py_ytd", "Prior Year YTD", py_fy_start, py_d),
        ("r12",    "Rolling 12M",    r12_start,   d),
        ("pr12",   "Prior 12M",      pr12_start,  pr12_end),
    ]


@frappe.whitelist()
def time_intelligence(config=None):
    """Aggregate chosen measures over standard analysis windows — MTD, prior
    MTD, QTD, YTD, prior-year YTD, rolling 12M, prior 12M — optionally split by
    a group field, with MoM% and YoY% deltas computed server-side.

    config = { doctype, date_field, measures:[fieldname...], group_by?,
               as_of?, filters:[...] }

    Field names are validated against the DocType meta before being embedded in
    aggregate expressions (no user text reaches SQL)."""
    _require_read()
    if isinstance(config, str):
        config = json.loads(config or "{}")
    doctype = config.get("doctype")
    if not doctype or not _can_read(doctype):
        frappe.throw(_("Pick a document you can read first."))
    meta = frappe.get_meta(doctype)
    ftype = {f.fieldname: f.fieldtype for f in meta.fields}
    flabel = {f.fieldname: (_(f.label) if f.label else f.fieldname) for f in meta.fields}

    date_field = config.get("date_field")
    if date_field not in ftype or ftype[date_field] not in ("Date", "Datetime"):
        frappe.throw(_("Pick a Date field for time intelligence."))
    measures = [m for m in (config.get("measures") or []) if ftype.get(m) in _NUMERIC][:6]
    if not measures:
        frappe.throw(_("Pick at least one numeric measure."))
    group_by = config.get("group_by") or None
    if group_by and group_by not in ftype:
        group_by = None

    from frappe.utils import nowdate
    as_of = config.get("as_of") or nowdate()
    base_filters = _coerce_filters(doctype, config.get("filters"))
    kwargs = {}
    if meta.istable:
        cp = _child_parent(doctype)
        if cp:
            base_filters = base_filters + [[doctype, "parenttype", "=", cp]]
            kwargs["parent_doctype"] = cp

    def flt_or_zero(v):
        try:
            return float(v or 0)
        except Exception:
            return 0.0

    windows = _ti_windows(as_of)
    agg_fields = ([group_by] if group_by else []) + [f"sum(`{m}`) as `{m}`" for m in measures]

    data = {}   # group_key -> window_key -> {measure: value}
    order = []  # stable group order (by first window encountered)
    for (wkey, _wlabel, ws, we) in windows:
        flt = base_filters + [[doctype, date_field, "between", [str(ws), str(we)]]]
        rows = frappe.get_list(doctype, filters=flt, fields=agg_fields,
                               group_by=group_by, limit_page_length=0, **kwargs)
        for r in rows:
            gkey = (r.get(group_by) or "—") if group_by else "__all__"
            if gkey not in data:
                data[gkey] = {}
                order.append(gkey)
            data[gkey][wkey] = {m: flt_or_zero(r.get(m)) for m in measures}

    def pct(cur, prev):
        return round((cur - prev) / abs(prev) * 100, 1) if abs(prev) > 1e-9 else None

    out_rows = []
    for gkey in order:
        w = data.get(gkey, {})
        row = {"group": gkey}
        for m in measures:
            g = lambda k: (w.get(k) or {}).get(m, 0.0)
            row[m] = {
                "mtd": round(g("mtd"), 2), "qtd": round(g("qtd"), 2),
                "ytd": round(g("ytd"), 2), "py_ytd": round(g("py_ytd"), 2),
                "r12": round(g("r12"), 2),
                "mom_pct": pct(g("mtd"), g("pmtd")),
                "yoy_pct": pct(g("ytd"), g("py_ytd")),
                "r12_pct": pct(g("r12"), g("pr12")),
            }
        out_rows.append(row)
    # Largest YTD of the first measure first — the natural reading order.
    out_rows.sort(key=lambda r: -abs(r[measures[0]]["ytd"]))

    return {
        "as_of": str(as_of),
        "windows": [{"key": k, "label": l, "from": str(s), "to": str(e)}
                    for (k, l, s, e) in windows],
        "group_by": group_by,
        "group_label": flabel.get(group_by) if group_by else None,
        "measures": [{"field": m, "label": flabel.get(m, m)} for m in measures],
        "rows": out_rows[:200],
    }


# ────────────────────────────────────────────────────────────────────────────
# Styled Excel export (v2.27.0) — the workbook mirrors the on-screen report:
# dark header, group bands, Sales/Returns bands, subtotal & grand-total
# styling, dd-mm-yyyy dates, thousands separators — plus a letterhead block
# (company name, address, VAT number) chosen by the user in the Studio print
# bar. Built server-side with openpyxl because client-side SheetJS cannot
# write styles.
# ────────────────────────────────────────────────────────────────────────────

@frappe.whitelist()
def export_xlsx(config=None, title=None, letter_head=None):
    _require_read()
    from datetime import date, datetime

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    if isinstance(config, str):
        config = json.loads(config or "{}")
    result = run_query(config)
    cols = result.get("columns") or []
    if not cols:
        frappe.throw(_("Run a report first — nothing to export."))

    wb = Workbook()
    ws = wb.active
    ws.title = (title or config.get("doctype") or "Report")[:31]
    ws.sheet_view.showGridLines = False

    # Palette mirrors the SPA report styling
    F_HEAD = PatternFill("solid", fgColor="2A2440")
    F_GRP = PatternFill("solid", fgColor="EFEAF9")
    F_BAND = PatternFill("solid", fgColor="F3F0FB")
    F_BANDR = PatternFill("solid", fgColor="FBEEF0")
    F_SUB = PatternFill("solid", fgColor="FAF7F0")
    F_SUBR = PatternFill("solid", fgColor="FDF2F4")
    F_NET = PatternFill("solid", fgColor="EEF7F4")
    thin = Side(style="thin", color="E6E0D4")
    B_ALL = Border(bottom=thin)
    B_TOP = Border(top=Side(style="medium", color="2A2440"))

    n_cols = len(cols)
    row_i = 1

    def _write(values, *, fill=None, bold=False, color=None, border=None,
               size=None, merge=False):
        nonlocal row_i
        if merge and n_cols > 1:
            ws.merge_cells(start_row=row_i, start_column=1, end_row=row_i, end_column=n_cols)
        for j, v in enumerate(values, start=1):
            c = ws.cell(row=row_i, column=j, value=v)
            if fill:
                c.fill = fill
            c.font = Font(bold=bold, color=color or "2A2440", size=size or 11)
            if border:
                c.border = border
        row_i += 1

    # Letterhead block — structured company fields (HTML can't render in xlsx)
    if letter_head is not None:
        try:
            from .report import get_letterhead
            lh = get_letterhead(letter_head or "", config.get("company"))
            if lh.get("company_name"):
                _write([lh["company_name"]], bold=True, size=15, merge=True)
            for line in (lh.get("address_lines") or [])[:3]:
                _write([line], color="777777", size=9, merge=True)
            contact = " · ".join(x for x in (lh.get("phone"), lh.get("email"),
                                             lh.get("website")) if x)
            if contact:
                _write([contact], color="777777", size=9, merge=True)
            if lh.get("tax_id"):
                _write([_("VAT: {0}").format(lh["tax_id"])], color="777777", size=9, merge=True)
            row_i += 1
        except Exception:
            pass

    _write([title or config.get("doctype") or _("Report")], bold=True, size=14, merge=True)
    sub = f"{config.get('doctype') or ''} · {result.get('row_count', 0)} {_('rows')}"
    if result.get("group_label"):
        sub += f" · {_('grouped by')} {result['group_label']}"
    sub += f" · {frappe.utils.formatdate(frappe.utils.nowdate(), 'dd-MM-yyyy')}"
    _write([sub], color="888888", size=9, merge=True)
    row_i += 1

    # Header row
    _write([c.get("label") or c["field"] for c in cols], fill=F_HEAD, bold=True, color="FFFFFF")

    date_cols = {j for j, c in enumerate(cols, start=1) if (c.get("type") in ("Date", "Datetime"))}
    num_cols_ix = {j for j, c in enumerate(cols, start=1) if c.get("numeric")}

    def _cell_val(c, r):
        v = r.get(c["field"])
        if c.get("type") in ("Date", "Datetime") and v:
            try:
                return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
            except Exception:
                return v
        if c.get("numeric"):
            try:
                return float(v or 0)
            except Exception:
                return v
        return v

    def _data_row(r):
        nonlocal row_i
        for j, c in enumerate(cols, start=1):
            cell = ws.cell(row=row_i, column=j, value=_cell_val(c, r))
            cell.border = B_ALL
            if j in date_cols:
                cell.number_format = "DD-MM-YYYY"
            elif j in num_cols_ix:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
        row_i += 1

    def _sub_row(label, sub, *, fill, color="8A6D3B", top=False):
        nonlocal row_i
        for j, c in enumerate(cols, start=1):
            v = label if j == 1 else (float(sub.get(c["field"]) or 0) if c.get("numeric") and c["field"] in (sub or {}) else None)
            cell = ws.cell(row=row_i, column=j, value=v)
            cell.fill = fill
            cell.font = Font(bold=True, color=color)
            if top:
                cell.border = B_TOP
            if c.get("numeric"):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
        row_i += 1

    if result.get("groups") is not None:
        for g in result["groups"]:
            _write([f"{result.get('group_label') or ''}: {g.get('key')} ({g.get('count')})"],
                   fill=F_GRP, bold=True, color="6C5CE7", merge=True)
            if result.get("split_returns"):
                _write([_("Sales")], fill=F_BAND, bold=True, color="6C5CE7", merge=True)
                for r in (g.get("sales_rows") or []):
                    _data_row(r)
                _sub_row(_("Sales Subtotal"), g.get("sales_subtotal") or {}, fill=F_SUB)
                if g.get("return_rows"):
                    _write([_("Returns")], fill=F_BANDR, bold=True, color="B3261E", merge=True)
                    for r in g["return_rows"]:
                        _data_row(r)
                    _sub_row(_("Returns Subtotal"), g.get("return_subtotal") or {}, fill=F_SUBR, color="B3261E")
                _sub_row(_("Net Subtotal"), g.get("net_subtotal") or {}, fill=F_NET, color="11816F")
            else:
                for r in (g.get("rows") or []):
                    _data_row(r)
                _sub_row(_("Subtotal"), g.get("subtotal") or {}, fill=F_SUB)
    else:
        for r in (result.get("rows") or []):
            _data_row(r)
    _sub_row(_("TOTAL"), result.get("grand_total") or {}, fill=F_NET, color="11816F", top=True)

    # Column widths from content sample
    for j, c in enumerate(cols, start=1):
        w = max(len(str(c.get("label") or "")), 10)
        for r in (result.get("rows") or [])[:50]:
            w = max(w, min(len(str(r.get(c["field"]) or "")), 42))
        ws.column_dimensions[get_column_letter(j)].width = w + 3

    import io
    buf = io.BytesIO()
    wb.save(buf)
    fname = frappe.scrub(title or config.get("doctype") or "report")
    frappe.local.response.filename = f"{fname}-{frappe.utils.nowdate()}.xlsx"
    frappe.local.response.filecontent = buf.getvalue()
    frappe.local.response.type = "binary"
